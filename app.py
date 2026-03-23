"""
Fersedo Production Management System
=====================================
Flask web application for managing production, procurement, employees, and vacations.
"""
# ─────────────────────────────────────────────────────────────
# IMPORTS
# ─────────────────────────────────────────────────────────────
import glob
import io
import json
import os
import random
import threading
import time
from io import BytesIO
from os import remove
from datetime import date, datetime, timedelta, timezone
from functools import wraps

import pandas as pd
import sqlite3
from argon2 import PasswordHasher, exceptions
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger
from flask import (
    Flask, flash, jsonify, redirect, render_template,
    request, send_file, send_from_directory, session, url_for
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch, mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    BaseDocTemplate, Frame, Image, PageBreak,
    PageTemplate, Paragraph, SimpleDocTemplate,
    Spacer, Table, TableStyle
)
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from werkzeug.utils import secure_filename
from PIL import Image as PILImage
from pywebpush import WebPushException, webpush
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from utils.config import DATABASE_PATH
from flask import send_from_directory

from extensions import db, ph, migrate
from config import Config
from utils.db import init_db, get_db
from utils.config import STATIC_FOLDER, POZICII_FOLDER, FONT_DIR, PROJECT_ROOT, PARTS_EXCEL, TEMPLATE_FOLDER
from utils.decorators import login_required, admin_required, worker_required


# ─────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────
VAPID_PUBLIC_KEY  = "BFeflCt8YzY6ctPC1CIYVe5V0cBSgn4JCOzSekSt0kKAlGna-ca2Bk0h9jS36xZc2M9w2QUrOXaKDKN14k9azxg"
VAPID_PRIVATE_KEY = "_9jeAqnzFVLfOOile0AyF9hDNYT_EkMskHxffOLoqlE"
VAPID_SUBJECT     = "mailto:nikolamartinkovic12@gmail.com"

AUTO_ASSIGN_INTERVAL_SECONDS = 14400  # 4 hours

EMAIL_HOST          = "smtp.gmail.com"
EMAIL_PORT          = 587
EMAIL_HOST_USER     = "fersedoo@gmail.com"
EMAIL_HOST_PASSWORD = "ejvu srce tvls wqtw"

# ─────────────────────────────────────────────────────────────
# APP INIT
# ─────────────────────────────────────────────────────────────
app = Flask(__name__,
    static_folder=STATIC_FOLDER,
    static_url_path="/static",
    template_folder=TEMPLATE_FOLDER,
)

init_db()

# ── Blueprints ──────────────────────────────────────────────
from routes.auth     import auth_bp
from routes.main     import main_bp
from routes.artikli  import artikli_bp
from routes.zalihi   import zalihi_bp
from routes.nabavki  import nabavki_bp
from routes.admin    import admin_bp
from routes.odmori   import odmori_bp
from routes.kvalitet import kvalitet_bp

app.register_blueprint(auth_bp)
app.register_blueprint(main_bp)
app.register_blueprint(artikli_bp)
app.register_blueprint(zalihi_bp)
app.register_blueprint(nabavki_bp)
app.register_blueprint(admin_bp)
app.register_blueprint(odmori_bp)
app.register_blueprint(kvalitet_bp)

app.config["DATABASE_PATH"] = DATABASE_PATH
app.config.from_object(Config)
db.init_app(app)
migrate.init_app(app, db)

app.secret_key = "твоја_многу_тајна_шифра_2025!@#"
app.config["VAPID_PUBLIC_KEY"]  = VAPID_PUBLIC_KEY
app.config["VAPID_PRIVATE_KEY"] = VAPID_PRIVATE_KEY
app.config["VAPID_SUBJECT"]     = VAPID_SUBJECT
app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{app.config['DATABASE_PATH']}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

assign_lock = threading.Lock()
last_deleted_records  = []
last_deleted_datum_od = None
last_deleted_datum_do = None
last_deleted_oddel    = None

# ─────────────────────────────────────────────────────────────
# FONT REGISTRATION
# ─────────────────────────────────────────────────────────────
def _register_fonts():
    for name, filename in [
        ("DejaVuSans",      "DejaVuSans.ttf"),
        ("DejaVuSans-Bold", "DejaVuSans-Bold.ttf"),
    ]:
        path = os.path.join(FONT_DIR, filename)
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(name, path))
            except Exception as e:
                print(f"[FONT] Could not register {name}: {e}")
        else:
            print(f"[FONT] File not found: {path}")

_register_fonts()

# ─────────────────────────────────────────────────────────────
# CONTEXT PROCESSORS & TEMPLATE FILTERS
# ─────────────────────────────────────────────────────────────
@app.context_processor
def inject_vapid_key():
    return dict(VAPID_PUBLIC_KEY=VAPID_PUBLIC_KEY)


@app.context_processor
def inject_modules():
    modules = {
        "basic": [
            {"value": "select_kamin",        "label": "Нов запис"},
            {"value": "add_part",            "label": "Отвори нов артикл"},
            {"value": "moj_zapisi",          "label": "Мои записи"},
            {"value": "kalkulacija",         "label": "Калкулација"},
            {"value": "artikli",             "label": "Артикли"},
            {"value": "nabavki",             "label": "Набавки"},
            {"value": "dashboard",           "label": "Dashboard"},
            {"value": "pregled_greski",      "label": "Грешки"},
            {"value": "admin_users",         "label": "Корисници"},
            {"value": "plan_proizvodstvo",   "label": "План за производство"},
            {"value": "izvestaj",            "label": "Извештај"},
            {"value": "procesni_cekori",     "label": "Процесни чекори"},
        ],
        "odmori": [
            {"value": "odmori",                      "label": "Одмори (главна)"},
            {"value": "baranje_odmor",               "label": "Барање за одмор"},
            {"value": "odmori_vraboteni",            "label": "Вработени"},
            {"value": "odmori_kalendar",             "label": "Календар"},
            {"value": "odmori_pregled_odmori",       "label": "Преглед на одмори"},
            {"value": "odmori_sekojdnevni_otsustva", "label": "Секојдневни отсуства"},
            {"value": "odmori_manager_emails",       "label": "📧 Email за менаџери"},
        ],
        "zalihi": [
            {"value": "zalihi",           "label": "Залихи"},
            {"value": "kvalitet",         "label": "Квалитет"},
            {"value": "email_recipients", "label": "📧 Email примачи"},
        ],
    }
    return dict(modules=modules)


@app.template_filter("datetimeformat")
def datetimeformat(value, format_string="%d/%m/%Y %H:%M"):
    if value is None:
        return ""
    if isinstance(value, str):
        try:
            dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            return value
    else:
        dt = value
    dt_local = dt.replace(tzinfo=timezone.utc) + timedelta(hours=1)
    return dt_local.strftime(format_string)

# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def save_part_to_excel(part_number, slika_filename):
    try:
        if not os.path.exists(PARTS_EXCEL):
            wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
            ws.append(["PartNumber", "Slika"]); wb.save(PARTS_EXCEL)
        df = pd.read_excel(PARTS_EXCEL, engine="openpyxl")
        mask = df["PartNumber"].astype(str) == str(part_number)
        if mask.any():
            df.loc[mask, "Slika"] = slika_filename
        else:
            df = pd.concat([df, pd.DataFrame([[part_number, slika_filename]],
                           columns=["PartNumber", "Slika"])], ignore_index=True)
        df.to_excel(PARTS_EXCEL, index=False, engine="openpyxl")
        return True
    except Exception as e:
        print(f"[EXCEL] Error saving part: {e}"); return False


def get_part_info(part_number):
    conn = get_db()
    row = conn.execute(
        "SELECT slika, ime FROM parts WHERE TRIM(part_number) = TRIM(?)", (part_number,)
    ).fetchone()
    conn.close()
    if row and row["slika"]:
        return {"ime": row["ime"], "slika": f"/static/parts/{row['slika']}"}
    return None


def add_page_number(canvas, doc):
    canvas.setFont("DejaVuSans", 9)
    canvas.setFillColor(colors.gray)
    canvas.drawRightString(190 * mm, 10 * mm,
        f"Страна {canvas.getPageNumber()} • {datetime.now().strftime('%d.%m.%Y %H:%M')}")


def get_compressed_image_buffer(slika_path, max_size=(450, 450), quality=68):
    try:
        img = PILImage.open(slika_path)
        img.thumbnail(max_size, PILImage.Resampling.LANCZOS)
        buf = BytesIO()
        img.convert("RGB").save(buf, format="JPEG", quality=quality, optimize=True)
        buf.seek(0)
        return buf
    except Exception as e:
        print(f"[IMAGE] Compression error for {slika_path}: {e}"); return None


def _format_cet(dt_str):
    try:
        dt_utc = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
        return (dt_utc + timedelta(hours=1)).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return dt_str


def _send_push(cursor, recipients, title, body, url="/nabavki"):
    payload = json.dumps({"title": title, "body": body, "icon": "/static/logo", "url": url})
    for recipient in recipients:
        username = recipient["username"]
        sub_row = cursor.execute(
            "SELECT subscription FROM push_subscriptions WHERE user = ?", (username,)
        ).fetchone()
        if not (sub_row and sub_row["subscription"]):
            continue
        try:
            webpush(subscription_info=json.loads(sub_row["subscription"]), data=payload,
                    vapid_private_key=VAPID_PRIVATE_KEY, vapid_claims={"sub": VAPID_SUBJECT})
            print(f"[PUSH] Sent to {username}")
        except WebPushException as ex:
            print(f"[PUSH] WebPushException for {username}: {ex.message}")
        except Exception as e:
            print(f"[PUSH] Error for {username}: {e}")


ALL_MODULES = (
    "select_kamin,add_part,moj_zapisi,kalkulacija,artikli,nabavki,dashboard,"
    "pregled_greski,admin_users,plan_proizvodstvo,izvestaj,procesni_cekori,"
    "odmori,baranje_odmor,odmori_vraboteni,odmori_kalendar,odmori_pregled_odmori,"
    "odmori_manager_emails,zalihi,kvalitet"
)

# ─────────────────────────────────────────────────────────────
# EMAIL / EXCEL / PDF HELPERS  (scheduler ги користи)
# ─────────────────────────────────────────────────────────────
def generiraj_zaliha_excel():
    try:
        conn = get_db()
        cursor = conn.cursor()
        rows = [dict(r) for r in cursor.execute("""
            SELECT p.part_number, p.ime, COALESCE(SUM(d.kolicina), 0) AS total
            FROM parts p LEFT JOIN zaliha_dodadi d ON p.id = d.artikl_id
            GROUP BY p.id HAVING total > 0 ORDER BY p.part_number
        """).fetchall()]
        conn.close()
        wb = Workbook(); ws = wb.active; ws.title = "Current Stock"
        logo_path = os.path.join(STATIC_FOLDER, "logo2.png")
        ws.merge_cells("A1:C7")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        if os.path.exists(logo_path):
            try:
                img = OpenpyxlImage(logo_path); img.width = 280; img.height = 110
                ws.add_image(img, "B1")
            except Exception as e:
                print(f"[LOGO ERROR] {e}")
        else:
            ws["A1"].value = "Fersedo"
            ws["A1"].font = Font(name="Calibri", bold=True, size=32, color="1E3A8A")
        ws.row_dimensions[8].height = 25
        title_row = 9
        ws.merge_cells(f"A{title_row}:C{title_row}")
        tc = ws[f"A{title_row}"]
        tc.value = f"CURRENT STOCK – {datetime.now().strftime('%d.%m.%Y')}"
        tc.font = Font(name="Calibri", bold=True, size=18, color="1E40AF")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[title_row].height = 40
        header_row = title_row + 2
        hfont = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        hfill = PatternFill("solid", fgColor="2563EB")
        for col_idx, header in enumerate(["Part Number", "Description", "Total Quantity"], 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = hfont; cell.fill = hfill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(**{s: Side(style="thin", color="D1D5DB") for s in ("left","right","top","bottom")})
        ws.row_dimensions[header_row].height = 36
        data_start = header_row + 1
        thin_border = Border(**{s: Side(style="thin", color="E5E7EB") for s in ("left","right","top","bottom")})
        for i, row in enumerate(rows):
            r = data_start + i
            fill = PatternFill("solid", fgColor="F9FAFB" if i % 2 == 0 else "FFFFFF")
            ws.cell(row=r, column=1).value = row["part_number"]
            ws.cell(row=r, column=2).value = row["ime"] or "Unknown item"
            ws.cell(row=r, column=3).value = row["total"]
            for col in range(1, 4):
                cell = ws.cell(row=r, column=col)
                cell.fill = fill; cell.border = thin_border
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(horizontal="center" if col == 3 else "left", vertical="center")
        total_row = data_start + len(rows) + 2
        ws.cell(row=total_row, column=1).value = "TOTAL"
        ws.cell(row=total_row, column=3).value = sum(r["total"] for r in rows)
        tfont = Font(name="Calibri", bold=True, size=14, color="065F46")
        tfill = PatternFill("solid", fgColor="D1FAE5")
        tborder = Border(**{s: Side(style="medium", color="10B981") for s in ("left","right","top","bottom")})
        for col in [1, 3]:
            cell = ws.cell(row=total_row, column=col)
            cell.font = tfont; cell.fill = tfill; cell.border = tborder
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[total_row].height = 42
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 20
        ws.freeze_panes = f"A{header_row + 1}"
        output = BytesIO(); wb.save(output); output.seek(0)
        print(f"[EMAIL ZALIHA] Excel generated — {len(rows)} items")
        return output
    except Exception as e:
        print(f"[EMAIL ZALIHA] Error: {e}")
        import traceback; traceback.print_exc()
        return None


def generiraj_zaliha_pdf():
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"],
                                 fontSize=18, textColor=colors.darkblue, spaceAfter=12)
    logo_path = os.path.join(STATIC_FOLDER, "logo2.png")
    if os.path.exists(logo_path):
        try:
            logo_img = Image(logo_path, width=120*mm, height=40*mm)
            logo_img.hAlign = "CENTER"
            elements.append(logo_img)
            elements.append(Spacer(1, 8*mm))
        except Exception as e:
            print(f"[PDF LOGO ERROR] {e}")
    elements.append(Paragraph(f"Current Stock Report – {datetime.now().strftime('%d.%m.%Y')}", title_style))
    elements.append(Spacer(1, 12*mm))
    table_data = [["Part Number", "Description", "Total Quantity"]]
    conn = get_db(); cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT p.part_number, p.ime, COALESCE(SUM(d.kolicina), 0) AS total
        FROM parts p LEFT JOIN zaliha_dodadi d ON p.id = d.artikl_id
        GROUP BY p.id HAVING total > 0 ORDER BY p.part_number
    """).fetchall()
    conn.close()
    total_qty = 0
    for row in rows:
        table_data.append([row["part_number"], row["ime"] or "Unknown item", row["total"]])
        total_qty += row["total"]
    table_data.append(["TOTAL", "", total_qty])
    table = Table(table_data, colWidths=[50*mm, 90*mm, 40*mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,0), 12),
        ("BOTTOMPADDING", (0,0), (-1,0), 12),
        ("BACKGROUND", (-1,-1), (-1,-1), colors.HexColor("#D1FAE5")),
        ("FONTNAME",   (-1,-1), (-1,-1), "Helvetica-Bold"),
        ("GRID",       (0,0), (-1,-1), 0.5, colors.lightgrey),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1), (-1,-2), [colors.white, colors.HexColor("#F9FAFB")]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 20*mm))
    elements.append(Paragraph(
        f"Generated by Fersedo Production System | {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        styles["Normal"]
    ))
    doc.build(elements)
    buffer.seek(0)
    return buffer


def isprati_zaliha_email():
    print(f"[EMAIL ZALIHA] Почнува — {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    with app.app_context():
        try:
            conn = get_db()
            rows = conn.execute("SELECT email, tip FROM email_recipients WHERE aktiven = 1").fetchall()
            conn.close()
            to_list = [r["email"] for r in rows if r["tip"] == "to"]
            cc_list = [r["email"] for r in rows if r["tip"] == "cc"]
            if not to_list:
                print("[EMAIL ZALIHA] Нема To примачи. Прескокнување."); return
            excel_buffer = generiraj_zaliha_excel()
            pdf_buffer   = generiraj_zaliha_pdf()
            if not excel_buffer or not pdf_buffer:
                print("[EMAIL ZALIHA] Неуспешно генерирање на прилог(и)"); return
            datum_str      = datetime.now().strftime("%d.%m.%Y")
            excel_filename = f"stock_report_{datum_str.replace('.', '')}.xlsx"
            pdf_filename   = f"stock_report_{datum_str.replace('.', '')}.pdf"
            msg = MIMEMultipart()
            msg["From"]    = EMAIL_HOST_USER
            msg["To"]      = ", ".join(to_list)
            if cc_list:
                msg["Cc"]  = ", ".join(cc_list)
            msg["Subject"] = f"📦 Current Stock Report – {datum_str}"
            msg.attach(MIMEText(
                f"Dear team,\nAttached are the current Fersedo stock reports as of {datum_str}.\n"
                f"This is an automatically generated report.\nBest regards,\nFersedo System",
                "plain", "utf-8"
            ))
            for buf, fname in [(excel_buffer, excel_filename), (pdf_buffer, pdf_filename)]:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(buf.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f'attachment; filename="{fname}"')
                msg.attach(part)
            with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT) as server:
                server.starttls()
                server.login(EMAIL_HOST_USER, EMAIL_HOST_PASSWORD)
                server.sendmail(EMAIL_HOST_USER, to_list + cc_list, msg.as_string())
            print(f"[EMAIL ZALIHA] Успешно испратено — To: {to_list}, Cc: {cc_list}")
        except Exception as e:
            print(f"[EMAIL ZALIHA] Грешка: {e}")
            import traceback; traceback.print_exc()

# ─────────────────────────────────────────────────────────────
# MISC ROUTES  (мали, не заслужуваат сопствен blueprint)
# ─────────────────────────────────────────────────────────────
@app.route("/api/part_info")
@login_required
def api_part_info():
    part_number = request.args.get("part_number", "").strip()
    if not part_number:
        return jsonify({"success": False, "error": "Нема part number"})
    result = get_part_info(part_number)
    if result and result.get("slika"):
        return jsonify({"success": True, "slika": result["slika"], "ime": result.get("ime", "")})
    return jsonify({"success": False, "error": "Не е пронајден"})


@app.route("/debug_routes")
def debug_routes():
    routes = [(r.endpoint, r.rule) for r in app.url_map.iter_rules()]
    return jsonify(sorted(routes))


@app.route("/upload_temp_image", methods=["POST"])
@login_required
def upload_temp_image():
    if "slika" not in request.files:
        return jsonify({"success": False, "error": "Нема датотека"}), 400
    file = request.files["slika"]
    if not file.filename:
        return jsonify({"success": False, "error": "Празно ime"}), 400
    ext = os.path.splitext(file.filename)[1].lower() or ".png"
    filename = f"camera_{int(time.time() * 1000)}{ext}"
    os.makedirs(POZICII_FOLDER, exist_ok=True)
    file.save(os.path.join(POZICII_FOLDER, filename))
    return jsonify({"success": True, "filename": filename})


@app.route("/subscribe", methods=["POST"])
@login_required
def subscribe():
    subscription = request.json
    conn = get_db(); cursor = conn.cursor()
    cursor.execute("DELETE FROM push_subscriptions WHERE user=?", (session["user"],))
    cursor.execute("INSERT INTO push_subscriptions (user, subscription) VALUES (?,?)",
                   (session["user"], json.dumps(subscription)))
    conn.commit(); conn.close()
    return jsonify({"success": True})


@app.route("/test_push")
@login_required
def test_push():
    conn = get_db()
    sub_row = conn.execute(
        "SELECT subscription FROM push_subscriptions WHERE user=?", (session["user"],)
    ).fetchone()
    conn.close()
    if sub_row:
        try:
            webpush(
                subscription_info=json.loads(sub_row["subscription"]),
                data=json.dumps({"title": "Тест известување", "body": "Тест push од Fersedo!", "url": "/nabavki"}),
                vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims={"sub": VAPID_SUBJECT},
            )
            return "Push испратен!"
        except Exception as e:
            return f"Грешка: {e}"
    return "Нема subscription"

@app.route('/manifest.json')
def manifest():
    return send_from_directory('static', 'manifest.json')

@app.route('/sw.js')
def sw():
    return send_from_directory('static', 'sw.js', mimetype='application/javascript')

# ─────────────────────────────────────────────────────────────
# AUTO-ASSIGN  (логика останува овде, scheduler исто)
# ─────────────────────────────────────────────────────────────
def auto_assign_nabavki():
    with assign_lock:
        try:
            conn = get_db(); cursor = conn.cursor()
            nabavki_users = [r["username"] for r in
                cursor.execute("SELECT username FROM users WHERE user_group='Nabavki'").fetchall()]
            if not nabavki_users:
                conn.close(); return
            pending = [r["id"] for r in cursor.execute("""
                SELECT id FROM nabavki_requests
                WHERE prevzemeno_od IS NULL AND status='креирано'
                ORDER BY datum_kreiranje ASC LIMIT 5
            """).fetchall()]
            assigned_count = 0
            for req_id in pending:
                chosen_user = min(
                    nabavki_users,
                    key=lambda u: (cursor.execute(
                        "SELECT COUNT(*) AS c FROM nabavki_requests WHERE prevzemeno_od=?", (u,)
                    ).fetchone() or {}).get("c", 0),
                    default=None,
                ) or random.choice(nabavki_users)
                now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                cursor.execute("""
                    UPDATE nabavki_requests
                    SET prevzemeno_od=?, datum_prevzemanje=?, status='Videno'
                    WHERE id=? AND prevzemeno_od IS NULL
                """, (chosen_user, now_str, req_id))
                if cursor.rowcount > 0:
                    cursor.execute("""
                        INSERT INTO nabavki_assign_log (request_id, assigned_to, assigned_at, from_thread)
                        VALUES (?,?,?,?)
                    """, (req_id, chosen_user, now_str, threading.current_thread().name))
                    conn.commit()
                    assigned_count += 1
                    print(f"[AUTO ASSIGN] {req_id} → {chosen_user}")
            print(f"[AUTO ASSIGN] {assigned_count} барања доделени." if assigned_count else "[AUTO ASSIGN] Нема pending.")
            conn.close()
        except Exception as e:
            print(f"[AUTO ASSIGN ERROR] {e}")


# ─────────────────────────────────────────────────────────────
# SCHEDULER
# ─────────────────────────────────────────────────────────────
from routes.odmori import isprati_dnevni_izvestaj_otsustva, isprati_nedelen_izvestaj_otsustva

scheduler = BackgroundScheduler()

scheduler.add_job(
    auto_assign_nabavki,
    trigger=IntervalTrigger(seconds=AUTO_ASSIGN_INTERVAL_SECONDS),
    id="nabavki_auto_assign",
    replace_existing=True,
)
scheduler.add_job(
    isprati_zaliha_email,
    trigger="cron",
    day_of_week="wed", hour=8, minute=0,
    id="zaliha_email_sreda",
    replace_existing=True,
)
scheduler.add_job(
    isprati_dnevni_izvestaj_otsustva,
    trigger="cron",
    day_of_week="mon-fri", hour=8, minute=0,
    id="otsustva_dnevni",
    replace_existing=True,
)
scheduler.add_job(
    isprati_nedelen_izvestaj_otsustva,
    trigger="cron",
    day_of_week="fri", hour=15, minute=0,
    id="otsustva_nedelen",
    replace_existing=True,
)

scheduler.start()
print("[SCHEDULER] Auto-assign started – interval 4 hours")
print("[SCHEDULER] Zaliha email started – every Wednesday at 08:00")
print("[SCHEDULER] Otsustva dnevni started – every day at 08:00")
print("[SCHEDULER] Otsustva nedelen started – every Friday at 15:00")

# ─────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=80,        # ← смени од 5001 на 80
        debug=False,
    )