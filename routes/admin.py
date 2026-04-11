# routes/admin.py
import sqlite3
import io
import os
import secrets
import time
import threading
from datetime import date, datetime, timedelta
from collections import defaultdict
from flask import Blueprint, current_app, render_template, request, flash, redirect, url_for, jsonify, send_file, session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, A5
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch, mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import BaseDocTemplate, Frame, Image, PageBreak, PageTemplate, Paragraph, Spacer, Table, TableStyle
from werkzeug.utils import secure_filename
from PIL import Image as PILImage
from utils.audit import get_audit_log, log_audit_event
from utils.backup_manager import (
    backup_path,
    create_backup,
    list_backups,
    restore_backup,
)
from utils.db import get_db
from utils.decorators import admin_or_module_required, admin_required, login_required
from argon2 import PasswordHasher, exceptions
from utils.config import POZICII_FOLDER, STATIC_FOLDER
from utils.helpers import add_page_number, get_compressed_image_buffer
from utils.odmori_helpers import get_email_log
from utils.runtime import read_app_log_entries
from utils.stock_reports import isprati_zaliha_email
from utils.emailing import send_new_user_credentials_email
from io import BytesIO

ph = PasswordHasher()


def _generate_temporary_password(length=10):
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz23456789"
    return "Fs-" + "".join(secrets.choice(alphabet) for _ in range(length))

last_deleted_records  = []
last_deleted_datum_od = None
last_deleted_datum_do = None
last_deleted_oddel    = None

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

# ─────────────────────────────────────────────────────────────
# USER MANAGEMENT
# ─────────────────────────────────────────────────────────────

ALL_MODULES = (
    "select_kamin,add_part,moj_zapisi,artikli,nabavki,nabavki_arhiva,"
    "dashboard,system_logs,pregled_greski,admin_users,email_recipients,izvestaj,procesni_cekori,"
    "odmori,baranje_odmor,odmori_vraboteni,odmori_kalendar,odmori_pregled_odmori,"
    "odmori_sekojdnevni_otsustva,odmori_manager_emails,zalihi,"
    "odrzuvanje,odrzuvanje_masini,odrzuvanje_nalozi,odrzuvanje_plan,odrzuvanje_istorija,"
    "kvalitet,kvalitet_nova,kvalitet_vlezna,kvalitet_arhiva,kvalitet_greski_statistika,kvalitet_template,"
    "ponudi,ponudi_arhiva,chat"
)

# Читливи имиња за секој модул (за приказ во admin_users.html)
MODULE_LABELS = {
    "select_kamin":               "Нов запис (производство)",
    "add_part":                   "Отвори нов артикл",
    "moj_zapisi":                 "Мои записи",
    "artikli":                    "Артикли",
    "nabavki":                    "Набавки",
    "nabavki_arhiva":             "Архива на набавки",
    "dashboard":                  "Dashboard",
    "system_logs":                "Системски логови",
    "pregled_greski":             "Преглед на грешки",
    "admin_users":                "Управување со корисници",
    "email_recipients":           "Email примачи",
    "izvestaj":                   "Извештај",
    "procesni_cekori":            "Процесни чекори",
    "odmori":                     "Одмори (главна)",
    "baranje_odmor":              "Барање за одмор",
    "odmori_vraboteni":           "Одмори — Вработени",
    "odmori_kalendar":            "Одмори — Календар",
    "odmori_pregled_odmori":      "Одмори — Преглед",
    "odmori_sekojdnevni_otsustva":"Одмори — Секојдневни отсуства",
    "odmori_manager_emails":      "Одмори — Email менаџери",
    "zalihi":                     "Залихи",
    "kvalitet":                   "Квалитет",
    "kvalitet_nova":              "Квалитет — Нова контрола",
    "kvalitet_vlezna":            "Квалитет — Влезна контрола",
    "kvalitet_arhiva":            "Квалитет — Архива",
    "kvalitet_greski_statistika": "Квалитет — Статистика на грешки",
    "kvalitet_template":          "Квалитет — QC Шаблони",
    "ponudi":                     "Понуди",
    "ponudi_arhiva":              "Архива на понуди",
    "chat":                       "Чат",
}


MODULE_LABELS.update(
    {
        "odrzuvanje": "Одржување",
        "odrzuvanje_masini": "Одржување — Машини",
        "odrzuvanje_nalozi": "Одржување — Налози",
        "odrzuvanje_plan": "Одржување — План",
        "odrzuvanje_istorija": "Одржување — Историја",
    }
)


@admin_bp.route("/users", methods=["GET", "POST"])
@login_required
@admin_or_module_required("admin_users")
def admin_users():
    conn   = get_db()
    cursor = conn.cursor()

    if request.method == "POST":
        action = request.form.get("action")

        if action == "create":
            username = request.form.get("username", "").strip()
            email = request.form.get("email", "").strip().lower()
            is_admin = int(request.form.get("is_admin", 0))
            user_group = request.form.get("user_group", "").strip()
            allowed_modules = ALL_MODULES if is_admin else ",".join(
                [m.strip() for m in request.form.getlist("allowed_modules") if m.strip()]
            )
            if not username or not email:
                flash("Корисничко име и email адреса се задолжителни!", "danger")
            else:
                temporary_password = _generate_temporary_password()
                try:
                    cursor.execute("""
                        INSERT INTO users
                        (username, hashed_password, is_admin, user_group, allowed_modules, email, must_change_password)
                        VALUES (?, ?, ?, ?, ?, ?, 1)
                    """, (username, ph.hash(temporary_password), is_admin, user_group, allowed_modules, email))
                    send_new_user_credentials_email(
                        email,
                        username,
                        temporary_password,
                        login_url=url_for("auth.login", _external=True),
                    )
                    conn.commit()
                    flash(
                        f'Корисникот <strong>{username}</strong> е успешно креиран и привремената лозинка е испратена на <strong>{email}</strong>.',
                        "success",
                    )
                except sqlite3.IntegrityError:
                    conn.rollback()
                    flash(f"Корисничкото име <strong>{username}</strong> веќе постои!", "danger")
                except Exception as e:
                    conn.rollback()
                    flash(f"Грешка при креирање: {str(e)}", "danger")

        elif action == "delete":
            username = request.form.get("username", "").strip()
            if username == session["user"]:
                flash("Не можете да го избришете сопствениот профил!", "warning")
            else:
                try:
                    cursor.execute("DELETE FROM users WHERE username = ?", (username,))
                    if cursor.rowcount > 0:
                        conn.commit()
                        flash(f"Корисникот <strong>{username}</strong> е успешно избришан!", "success")
                    else:
                        flash(f"Корисникот <strong>{username}</strong> не постои.", "warning")
                except Exception as e:
                    flash(f"Грешка при бришење: {str(e)}", "danger")

        elif action == "edit":
            username = request.form.get("username", "").strip()
            is_admin = int(request.form.get("is_admin", 0))
            user_group = request.form.get("user_group", "").strip()
            email = request.form.get("email", "").strip().lower()
            allowed_modules = ALL_MODULES if is_admin else ",".join(
                [m.strip() for m in request.form.getlist("allowed_modules") if m.strip()]
            )
            try:
                cursor.execute("""
                    UPDATE users
                    SET is_admin=?, user_group=?, allowed_modules=?, email=?
                    WHERE username=?
                """, (is_admin, user_group, allowed_modules, email, username))
                if cursor.rowcount > 0:
                    conn.commit()
                    flash(f"Корисникот <strong>{username}</strong> е успешно ажуриран!", "success")
                else:
                    flash(f"Корисникот <strong>{username}</strong> не постои.", "warning")
            except Exception as e:
                conn.rollback()
                flash(f"Грешка при ажурирање: {str(e)}", "danger")

        conn.close()
        return redirect(url_for("admin.admin_users"))

    users = cursor.execute("""
        SELECT username, is_admin, user_group, allowed_modules, COALESCE(email, '') AS email
        FROM users ORDER BY username
    """).fetchall()
    conn.close()
    modules = {
        "basic": [
            {"value": "select_kamin",      "label": "Нов запис"},
            {"value": "add_part",          "label": "Отвори нов артикл"},
            {"value": "moj_zapisi",        "label": "Мои записи"},
            {"value": "artikli",           "label": "Артикли"},
            {"value": "dashboard",         "label": "Dashboard"},
            {"value": "system_logs",       "label": "Системски логови"},
            {"value": "pregled_greski",    "label": "Грешки"},
            {"value": "admin_users",       "label": "Корисници"},
            {"value": "izvestaj",          "label": "Извештај"},
            {"value": "procesni_cekori",   "label": "Процесни чекори"},
        ],
        "nabavki": [
            {"value": "nabavki",           "label": "Набавки"},
            {"value": "nabavki_arhiva",    "label": "Архива на набавки"},
            {"value": "ponudi",            "label": "Понуди"},
            {"value": "ponudi_arhiva",     "label": "Архива на понуди"},
            {"value": "chat",              "label": "Чат"},
        ],
        "odmori": [
            {"value": "odmori",                      "label": "Одмори (главна)"},
            {"value": "baranje_odmor",               "label": "Барање за одмор"},
            {"value": "odmori_vraboteni",            "label": "Вработени"},
            {"value": "odmori_kalendar",             "label": "Календар"},
            {"value": "odmori_pregled_odmori",       "label": "Преглед на одмори"},
            {"value": "odmori_sekojdnevni_otsustva", "label": "Секојдневни отсуства"},
            {"value": "odmori_manager_emails",       "label": "Email за менаџери"},
        ],
        "zalihi": [
            {"value": "zalihi", "label": "Залихи"},
            {"value": "email_recipients", "label": "Email примачи"},
        ],
        "odrzuvanje": [
            {"value": "odrzuvanje", "label": "Одржување"},
            {"value": "odrzuvanje_masini", "label": "Машини"},
            {"value": "odrzuvanje_nalozi", "label": "Налози"},
            {"value": "odrzuvanje_plan", "label": "План"},
            {"value": "odrzuvanje_istorija", "label": "Историја"},
        ],
        "kvalitet": [
            {"value": "kvalitet",          "label": "Квалитет"},
            {"value": "kvalitet_nova",     "label": "Нова контрола"},
            {"value": "kvalitet_vlezna",   "label": "Влезна контрола"},
            {"value": "kvalitet_arhiva",   "label": "Архива на контроли"},
            {"value": "kvalitet_greski_statistika", "label": "Статистика на грешки"},
            {"value": "kvalitet_template", "label": "QC Шаблони"},
        ],
    }
    return render_template("admin_users.html",
                           users=users,
                           current_user=session["user"],
                           modules=modules)


# ─────────────────────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────────────────────
@admin_bp.route("/dashboard")
@login_required
@admin_or_module_required("dashboard")
def dashboard():
    return render_template("dashboard.html", today=date.today().isoformat())


@admin_bp.route("/backup")
@login_required
@admin_required
def admin_backup():
    backups = list_backups(limit=200)
    return render_template(
        "admin_backup.html",
        backups=backups,
        auto_backup_enabled=current_app.config.get("AUTO_BACKUP_ENABLED", True),
        auto_backup_hour=int(current_app.config.get("AUTO_BACKUP_HOUR", 2)),
        auto_backup_minute=int(current_app.config.get("AUTO_BACKUP_MINUTE", 30)),
        backup_keep=int(current_app.config.get("AUTO_BACKUP_KEEP", 30)),
        backup_dir=current_app.config.get("BACKUP_DIR", ""),
    )


@admin_bp.route("/backup/create", methods=["POST"])
@login_required
@admin_required
def admin_backup_create():
    try:
        created = create_backup(reason=f"manual_{session.get('user', 'admin')}")
        log_audit_event(
            "admin",
            "backup_create",
            username=session.get("user", ""),
            status="success",
            details=f"Created backup: {created['name']}",
        )
        flash(f"Backup e kreiran: {created['name']}", "success")
    except Exception as exc:
        log_audit_event(
            "admin",
            "backup_create",
            username=session.get("user", ""),
            status="error",
            details=f"Create backup failed: {exc}",
        )
        flash(f"Greska pri kreiranje backup: {exc}", "danger")
    return redirect(url_for("admin.admin_backup"))


@admin_bp.route("/backup/restore", methods=["POST"])
@login_required
@admin_required
def admin_backup_restore():
    backup_name = (request.form.get("backup_name") or "").strip()
    if not backup_name:
        flash("Izberi backup za restore.", "warning")
        return redirect(url_for("admin.admin_backup"))

    try:
        restored = restore_backup(backup_name)
        log_audit_event(
            "admin",
            "backup_restore",
            username=session.get("user", ""),
            status="warning",
            details=(
                f"Restored from {restored['restored_from']} "
                f"(safety backup: {restored['safety_backup']})"
            ),
        )
        flash(
            f"Restore uspesno. Vrateno od {restored['restored_from']} "
            f"(bezbednosen backup: {restored['safety_backup']}).",
            "success",
        )
    except Exception as exc:
        log_audit_event(
            "admin",
            "backup_restore",
            username=session.get("user", ""),
            status="error",
            details=f"Restore failed for {backup_name}: {exc}",
        )
        flash(f"Greska pri restore: {exc}", "danger")
    return redirect(url_for("admin.admin_backup"))


@admin_bp.route("/backup/download/<path:backup_name>")
@login_required
@admin_required
def admin_backup_download(backup_name):
    try:
        path = backup_path(backup_name)
        return send_file(path, as_attachment=True, download_name=path.name, max_age=0)
    except Exception:
        flash("Backup datotekata ne e pronajdena.", "warning")
        return redirect(url_for("admin.admin_backup"))


@admin_bp.route("/system_logs")
@login_required
@admin_or_module_required("system_logs")
def system_logs():
    log_audit_event(
        "admin",
        "system_logs_view",
        status="info",
        details="Отворен е прегледот за системски логови",
    )
    runtime_log = read_app_log_entries(current_app, limit=80)
    audit_log = get_audit_log(limit=40)
    email_log = get_email_log(limit=25)

    error_count = sum(1 for entry in runtime_log if entry.get("level") in {"ERROR", "CRITICAL"})
    warning_count = sum(1 for entry in runtime_log if entry.get("level") == "WARNING")
    info_count = sum(1 for entry in runtime_log if entry.get("level") == "INFO")

    return render_template(
        "system_logs.html",
        runtime_log=runtime_log,
        audit_log=audit_log,
        email_log=email_log,
        error_count=error_count,
        warning_count=warning_count,
        info_count=info_count,
    )


@admin_bp.route("/api/greski")
@login_required
@admin_or_module_required("pregled_greski")
def api_greski():
    datum_od = request.args.get("datum_od")
    datum_do = request.args.get("datum_do")
    if not datum_od or not datum_do:
        return jsonify({"error": "Датуми се задолжителни"}), 400
    try:
        datetime.strptime(datum_od, "%Y-%m-%d")
        datum_do_plus = (datetime.strptime(datum_do, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
        conn   = get_db()
        cursor = conn.cursor()
        rows = cursor.execute("""
            SELECT oddel, SUM(greski) AS total_greski, SUM(proizvedeni) AS total_proizvedeni
            FROM performance
            WHERE datum >= ? AND datum < ? AND tip_proizvod = 'polugotov'
            GROUP BY oddel
        """, (datum_od, datum_do_plus)).fetchall()
        total_row = cursor.execute("""
            SELECT SUM(greski) AS total_greski_all, SUM(proizvedeni) AS total_proizvedeni_all
            FROM performance
            WHERE datum >= ? AND datum < ? AND tip_proizvod = 'polugotov'
        """, (datum_od, datum_do_plus)).fetchone()
        conn.close()
        data = []
        for row in rows:
            g = row["total_greski"] or 0
            p = row["total_proizvedeni"] or 0
            efekt = round(((p - g) / p) * 100, 1) if p > 0 else 0
            data.append({"oddel": row["oddel"], "greski": g, "proizvedeni": p, "efektivnost": efekt})
        g_all       = total_row["total_greski_all"] or 0
        p_all       = total_row["total_proizvedeni_all"] or 0
        total_efekt = round(((p_all - g_all) / p_all) * 100, 1) if p_all > 0 else 0
        return jsonify({"per_oddel": data, "total": {"greski": g_all, "proizvedeni": p_all, "efektivnost": total_efekt}})
    except ValueError as e:
        return jsonify({"error": f"Невалиден формат на датум: {e}"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────────────────────
# PRODUCTION ERRORS (GRESKI)
# ─────────────────────────────────────────────────────────────
@admin_bp.route("/greski", methods=["GET", "POST"])
@login_required
@admin_or_module_required("pregled_greski")
def pregled_greski():
    global last_deleted_records, last_deleted_datum_od, last_deleted_datum_do, last_deleted_oddel
    datum_od = (request.form if request.method == "POST" else request.args).get("datum_od")
    datum_do = (request.form if request.method == "POST" else request.args).get("datum_do")
    if not datum_od or not datum_do:
        datum_od = (date.today() - timedelta(days=30)).isoformat()
        datum_do = date.today().isoformat()
    conn          = get_db()
    datum_do_plus = (datetime.strptime(datum_do, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    if request.method == "POST":
        action = request.form.get("action")
        if action == "delete":
            oddel          = request.form.get("oddel")
            admin_password = request.form.get("admin_password")
            user = conn.execute("SELECT hashed_password FROM users WHERE username = ?", (session["user"],)).fetchone()
            try:
                ph.verify(user["hashed_password"], admin_password)
            except exceptions.VerifyMismatchError:
                flash("Погрешна лозинка!", "error")
                conn.close()
                return redirect(url_for("admin.pregled_greski", datum_od=datum_od, datum_do=datum_do))
            last_deleted_records = conn.execute("""
                SELECT * FROM performance
                WHERE datum >= ? AND datum < ? AND tip_proizvod = 'polugotov' AND (oddel = ? OR ? IS NULL)
            """, (datum_od, datum_do_plus, oddel, oddel)).fetchall()
            last_deleted_datum_od = datum_od
            last_deleted_datum_do = datum_do
            last_deleted_oddel    = oddel
            cursor = conn.cursor()
            cursor.execute("""
                DELETE FROM performance
                WHERE datum >= ? AND datum < ? AND tip_proizvod = 'polugotov' AND (oddel = ? OR ? IS NULL)
            """, (datum_od, datum_do_plus, oddel, oddel))
            conn.commit()
            flash(f"Успешно избришани {cursor.rowcount} записи!", "success")
        elif action == "undo":
            if not last_deleted_records:
                flash("Нема што да се врати!", "error")
            else:
                cursor = conn.cursor()
                for rec in last_deleted_records:
                    cursor.execute("""
                        INSERT INTO performance
                        (datum, oddel, proizvod, proizvedeni, greski, vid_greska, zabeleska,
                         username, timestamp, kamin, part_number, tip_proizvod)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (rec["datum"], rec["oddel"], rec["proizvod"], rec["proizvedeni"],
                          rec["greski"], rec["vid_greska"], rec["zabeleska"], rec["username"],
                          rec["timestamp"], rec["kamin"], rec["part_number"], "polugotov"))
                conn.commit()
                flash("Последното бришење е успешно вратено!", "success")
        conn.close()
        return redirect(url_for("admin.pregled_greski", datum_od=datum_od, datum_do=datum_do))
    records = conn.execute("""
        SELECT datum, timestamp, oddel, proizvod, greski, vid_greska, zabeleska,
               username, kamin, part_number, proizvedeni
        FROM performance
        WHERE datum >= ? AND datum < ? AND tip_proizvod = 'polugotov'
        ORDER BY timestamp DESC
    """, (datum_od, datum_do_plus)).fetchall()
    conn.close()
    return render_template("pregled_greski.html", records=records, datum_od=datum_od,
                           datum_do=datum_do, has_undo=bool(last_deleted_records))


@admin_bp.route("/greski/export")
@login_required
@admin_or_module_required("pregled_greski")
def export_greski():
    datum_od = request.args.get("datum_od")
    datum_do = request.args.get("datum_do")
    if not datum_od or not datum_do:
        flash("Мора да изберете период!", "error")
        return redirect(url_for("admin.pregled_greski"))
    datum_do_plus = (datetime.strptime(datum_do, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    conn    = get_db()
    records = conn.execute("""
        SELECT datum, timestamp, oddel, proizvod, greski, vid_greska, zabeleska,
               username, kamin, part_number, proizvedeni
        FROM performance
        WHERE datum >= ? AND datum < ? AND tip_proizvod = 'polugotov'
        ORDER BY timestamp DESC
    """, (datum_od, datum_do_plus)).fetchall()
    conn.close()
    wb = Workbook(); ws = wb.active; ws.title = "Грешки"
    ws.append(["Датум","Време","Оддел","Производ","Грешки","Произведени",
               "Вид грешка","Забелешка","Корисник","Камин","Part number"])
    for r in records:
        ws.append([r["datum"],r["timestamp"],r["oddel"],r["proizvod"],
                   r["greski"],r["proizvedeni"] or 0,r["vid_greska"] or "-",
                   r["zabeleska"] or "-",r["username"] or "Непознат",
                   r["kamin"] or "-",r["part_number"] or "-"])
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=f"greski_{datum_od}_do_{datum_do}.xlsx")


# ─────────────────────────────────────────────────────────────
# PROCESNI CEKORI
# ─────────────────────────────────────────────────────────────
@admin_bp.route("/procesni_cekori")
@login_required
@admin_or_module_required("procesni_cekori")
def procesni_cekori():
    conn   = get_db()
    kamini = [r["ime"] for r in conn.execute("SELECT ime FROM kamini ORDER BY ime").fetchall()]
    conn.close()
    return render_template("procesni_cekori.html", kamini=kamini)


@admin_bp.route("/procesni_cekori/<kamin>", methods=["GET", "POST"])
@login_required
@admin_or_module_required("procesni_cekori")
def pozicii_kamin(kamin):
    conn   = get_db()
    cursor = conn.cursor()
    if request.method == "POST":
        try:
            pozicija_broj_str = request.form.get("pozicija_broj", "").strip()
            if not pozicija_broj_str.isdigit() or int(pozicija_broj_str) < 1:
                flash("Број на позиција мора да биде валиден позитивен број!", "danger")
                conn.close()
                return redirect(url_for("admin.pozicii_kamin", kamin=kamin))
            pozicija_broj  = int(pozicija_broj_str)
            part_number    = request.form.get("part_number", "").strip() or None
            slika_filename = None
            if "slika" in request.files and request.files["slika"].filename:
                file = request.files["slika"]
                ext  = os.path.splitext(file.filename)[1].lower() or ".png"
                slika_filename = f"{secure_filename(kamin)}_{pozicija_broj}_{int(time.time() * 1000)}{ext}"
                os.makedirs(POZICII_FOLDER, exist_ok=True)
                file.save(os.path.join(POZICII_FOLDER, slika_filename))
            elif camera_filename := request.form.get("camera_slika_filename"):
                slika_filename = camera_filename.strip()
                if not os.path.exists(os.path.join(POZICII_FOLDER, slika_filename)):
                    flash("Сликата од камерата не е пронајдена на серверот!", "warning")
            cursor.execute("""
                INSERT INTO procesni_pozicii (kamin, pozicija_broj, part_number, slika)
                VALUES (?, ?, ?, ?)
            """, (kamin, pozicija_broj, part_number, slika_filename))
            conn.commit()
            flash(f"Позиција {pozicija_broj} е успешно додадена!", "success")
        except Exception as e:
            flash(f"Грешка при додавање позиција: {e}", "danger")
            conn.rollback()
        finally:
            conn.close()
            return redirect(url_for("admin.pozicii_kamin", kamin=kamin))
    pozicii = cursor.execute("""
        SELECT id, pozicija_broj, part_number, slika
        FROM procesni_pozicii WHERE kamin=? ORDER BY pozicija_broj ASC, id ASC
    """, (kamin,)).fetchall()
    conn.close()
    return render_template("pozicii_kamin.html", kamin=kamin, pozicii=[dict(p) for p in pozicii])


@admin_bp.route("/procesni_cekori/<kamin>/delete/<int:pozicija_id>", methods=["POST"])
@login_required
@admin_or_module_required("procesni_cekori")
def delete_pozicija(kamin, pozicija_id):
    conn   = get_db()
    cursor = conn.cursor()
    try:
        row = cursor.execute("SELECT slika FROM procesni_pozicii WHERE id=? AND kamin=?", (pozicija_id, kamin)).fetchone()
        if row and row["slika"]:
            path = os.path.join(POZICII_FOLDER, row["slika"])
            if os.path.exists(path):
                try: os.remove(path)
                except Exception as e: print(f"[FILE] Could not delete: {e}")
        cursor.execute("DELETE FROM procesni_pozicii WHERE id=? AND kamin=?", (pozicija_id, kamin))
        conn.commit()
        flash("Позицијата е успешно избришана!" if cursor.rowcount else "Позицијата не е пронајдена.",
              "success" if cursor.rowcount else "error")
    except Exception as e:
        flash(f"Грешка при бришење: {e}", "error"); conn.rollback()
    finally:
        conn.close()
    return redirect(url_for("admin.pozicii_kamin", kamin=kamin))


@admin_bp.route("/procesni_cekori/<kamin>/edit/<int:pozicija_id>", methods=["GET", "POST"])
@login_required
@admin_or_module_required("procesni_cekori")
def edit_pozicija(kamin, pozicija_id):
    conn     = get_db()
    cursor   = conn.cursor()
    pozicija = cursor.execute(
        "SELECT id, pozicija_broj, part_number, slika FROM procesni_pozicii WHERE id=? AND kamin=?",
        (pozicija_id, kamin)
    ).fetchone()
    if not pozicija:
        flash("Позицијата не е пронајдена!", "error")
        conn.close()
        return redirect(url_for("admin.pozicii_kamin", kamin=kamin))
    pozicija = dict(pozicija)
    if request.method == "POST":
        try:
            nov_broj   = int(request.form.get("pozicija_broj", 0))
            nov_part   = request.form.get("part_number", "").strip()
            slika_file = request.files.get("slika")
            slika_filename = pozicija["slika"]
            if nov_broj < 1:
                flash("Број на позиција мора да биде ≥ 1!", "error")
                conn.close()
                return redirect(url_for("admin.edit_pozicija", kamin=kamin, pozicija_id=pozicija_id))
            if slika_file and slika_file.filename:
                ext  = os.path.splitext(slika_file.filename)[1].lower() or ".png"
                pn_safe = nov_part or "bez_part"
                slika_filename = f"{secure_filename(kamin)}_{nov_broj}_{secure_filename(pn_safe)}_{int(time.time() * 1000)}{ext}"
                slika_file.save(os.path.join(POZICII_FOLDER, slika_filename))
                if pozicija["slika"] and pozicija["slika"] != slika_filename:
                    old_path = os.path.join(POZICII_FOLDER, pozicija["slika"])
                    if os.path.exists(old_path):
                        try: os.remove(old_path)
                        except Exception as e: print(f"[FILE] Could not delete old image: {e}")
            cursor.execute("""
                UPDATE procesni_pozicii SET pozicija_broj=?, part_number=?, slika=? WHERE id=? AND kamin=?
            """, (nov_broj, nov_part, slika_filename, pozicija_id, kamin))
            conn.commit()
            flash("Позицијата е успешно ажурирана!", "success")
        except ValueError:
            flash("Број на позиција мора да биде валиден број!", "error")
        except Exception as e:
            flash(f"Грешка: {e}", "error")
        conn.close()
        return redirect(url_for("admin.pozicii_kamin", kamin=kamin))
    conn.close()
    return render_template("edit_pozicija.html", kamin=kamin, pozicija=pozicija)


@admin_bp.route("/procesni_cekori/<kamin>/export")
@login_required
@admin_or_module_required("procesni_cekori")
def export_procesen_cekor(kamin):
    conn    = get_db()
    pozicii = conn.execute("""
        SELECT pozicija_broj, part_number, slika FROM procesni_pozicii WHERE kamin=? ORDER BY pozicija_broj ASC
    """, (kamin,)).fetchall()
    conn.close()
    if not pozicii:
        flash(f"Нема позиции за {kamin}!", "error")
        return redirect(url_for("admin.pozicii_kamin", kamin=kamin))
    wb = Workbook(); ws = wb.active; ws.title = f"Procesen Cekor - {kamin}"
    ws.append(["Позиција","Part Number","Слика"])
    for p in pozicii:
        ws.append([p["pozicija_broj"], p["part_number"], p["slika"] or "Нема слика"])
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=f"procesen_cekor_{kamin.replace(' ', '_')}.xlsx")


# ─────────────────────────────────────────────────────────────
# EMAIL RECIPIENTS (ЗАЛИХИ)
# ─────────────────────────────────────────────────────────────
@admin_bp.route("/email_recipients", methods=["GET", "POST"])
@login_required
@admin_or_module_required("email_recipients")
def email_recipients():
    conn   = get_db()
    cursor = conn.cursor()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            email = request.form.get("email", "").strip().lower()
            tip   = request.form.get("tip", "to")
            if email:
                try:
                    cursor.execute("""
                        INSERT INTO email_recipients (email, tip, aktiven, dodaden_od)
                        VALUES (?, ?, 1, ?)
                    """, (email, tip, session["user"]))
                    conn.commit()
                    flash(f"✅ Додаден: {email} ({tip.upper()})", "success")
                except sqlite3.IntegrityError:
                    flash(f"⚠️ {email} веќе постои!", "warning")
                except Exception as e:
                    flash(f"Грешка: {e}", "danger")
        elif action == "toggle":
            rec_id = request.form.get("rec_id")
            cursor.execute("UPDATE email_recipients SET aktiven = 1 - aktiven WHERE id = ?", (rec_id,))
            conn.commit(); flash("Статусот е променет!", "success")
        elif action == "delete":
            rec_id = request.form.get("rec_id")
            cursor.execute("DELETE FROM email_recipients WHERE id = ?", (rec_id,))
            conn.commit(); flash("Примачот е избришан!", "success")
        elif action == "change_tip":
            rec_id  = request.form.get("rec_id")
            nov_tip = request.form.get("nov_tip")
            cursor.execute("UPDATE email_recipients SET tip = ? WHERE id = ?", (nov_tip, rec_id))
            conn.commit(); flash("Типот е изменет!", "success")
        conn.close()
        return redirect(url_for("admin.email_recipients"))
    recipients = cursor.execute(
        "SELECT * FROM email_recipients ORDER BY tip ASC, email ASC"
    ).fetchall()
    conn.close()
    return render_template("email_recipients.html", recipients=recipients)


@admin_bp.route("/test_zaliha_email")
@login_required
@admin_or_module_required("email_recipients")
def test_zaliha_email():
    try:
        isprati_zaliha_email()
        flash("✅ Email за залиха е успешно испратен!", "success")
    except Exception as e:
        flash(f"❌ Грешка: {str(e)}", "danger")
    return redirect(url_for("admin.email_recipients"))


@admin_bp.route("/auto_assign_status")
@login_required
@admin_required
def auto_assign_status():
    conn   = get_db()
    cursor = conn.cursor()
    nabavki_users = cursor.execute("SELECT username FROM users WHERE user_group='Nabavki'").fetchall()
    free_requests = cursor.execute("""
        SELECT COUNT(*) AS count FROM nabavki_requests
        WHERE status='креирано' AND (prevzemeno_od IS NULL OR prevzemeno_od='')
    """).fetchone()["count"]
    user_stats = [{
        "username": u["username"],
        "assigned": cursor.execute(
            "SELECT COUNT(*) AS c FROM nabavki_requests WHERE prevzemeno_od=?", (u["username"],)
        ).fetchone()["c"],
    } for u in nabavki_users]
    try:
        last_assignments = cursor.execute(
            "SELECT * FROM nabavki_assign_log ORDER BY assigned_at DESC LIMIT 5"
        ).fetchall()
    except Exception:
        last_assignments = []
    conn.close()
    return render_template("auto_assign_status.html",
                           nabavki_users=user_stats,
                           free_requests=free_requests,
                           last_assignments=last_assignments,
                           current_time=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


@admin_bp.route("/test_auto_assign")
@login_required
@admin_required
def test_auto_assign():
    try:
        conn   = get_db()
        cursor = conn.cursor()
        users   = cursor.execute("SELECT username FROM users WHERE user_group='Nabavki'").fetchall()
        free    = cursor.execute("""
            SELECT COUNT(*) AS free FROM nabavki_requests
            WHERE status='креирано' AND (prevzemeno_od IS NULL OR prevzemeno_od='')
        """).fetchone()["free"]
        examples = cursor.execute("""
            SELECT id, naslov, username FROM nabavki_requests
            WHERE status='креирано' AND (prevzemeno_od IS NULL OR prevzemeno_od='')
            LIMIT 5
        """).fetchall()
        conn.close()
        return jsonify({
            "status": "ok",
            "threads": [t.name for t in threading.enumerate()],
            "nabavki_users": [u["username"] for u in users],
            "free_requests": free,
            "example_requests": [dict(e) for e in examples],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@admin_bp.route("/procesni_cekori/<kamin>/pdf")
@login_required
@admin_or_module_required("procesni_cekori")
def export_procesen_cekor_pdf(kamin):
    conn = get_db()
    rows = conn.execute("""
        SELECT pozicija_broj, part_number, slika FROM procesni_pozicii
        WHERE kamin=? ORDER BY pozicija_broj ASC, id ASC
    """, (kamin,)).fetchall()
    conn.close()
    if not rows:
        flash(f"Нема позиции за {kamin}!", "error")
        return redirect(url_for("admin.pozicii_kamin", kamin=kamin))
    buffer   = BytesIO()
    doc      = BaseDocTemplate(buffer, pagesize=A5, rightMargin=10*mm, leftMargin=10*mm,
                               topMargin=8*mm, bottomMargin=16*mm)
    frame    = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="normal")
    template = PageTemplate(id="All", frames=frame, onPage=add_page_number)
    doc.addPageTemplates([template])
    title_style   = ParagraphStyle("Title", fontName="DejaVuSans-Bold", fontSize=17, leading=20,
                                   textColor=colors.HexColor("#111827"), alignment=1, spaceAfter=10, spaceBefore=2)
    heading_style = ParagraphStyle("Heading", fontName="DejaVuSans-Bold", fontSize=11, leading=13,
                                   textColor=colors.HexColor("#1e40af"), spaceAfter=8, spaceBefore=10)
    grouped = defaultdict(list)
    for row in rows:
        grouped[row["pozicija_broj"]].append(row)
    elements   = []
    first_page = True
    for poz_broj in sorted(grouped.keys()):
        if not first_page:
            elements.append(PageBreak())
        first_page = False
        logo_path = os.path.join(STATIC_FOLDER, "logo.webp")
        if os.path.exists(logo_path):
            try:
                with open(logo_path, "rb") as f:
                    logo_buf = BytesIO(f.read())
                logo       = Image(logo_buf, width=1.05*inch, height=0.4*inch)
                logo.hAlign = "LEFT"
                elements.append(logo)
            except Exception:
                pass
        elements.append(Paragraph(f"Процесни чекори за камин: {kamin}", title_style))
        elements.append(Spacer(1, 0.12*inch))
        count = len(grouped[poz_broj])
        elements.append(Paragraph(f"Позиција {poz_broj} ({count} елемент{'и' if count > 1 else ''})", heading_style))
        elements.append(Spacer(1, 0.08*inch))
        data = [["Part Number", "Слика"]]
        for p in grouped[poz_broj]:
            slika_cell = "Нема слика"
            if p["slika"]:
                slika_path = os.path.join(POZICII_FOLDER, p["slika"])
                if os.path.exists(slika_path):
                    img_buf = get_compressed_image_buffer(slika_path, max_size=(240, 240), quality=48)
                    if img_buf:
                        try:
                            img        = Image(img_buf, width=0.95*inch, height=0.72*inch, kind="direct")
                            img.hAlign = "CENTER"
                            slika_cell = img
                        except Exception:
                            slika_cell = "Грешка"
            data.append([p["part_number"] or "— (без артикл)", slika_cell])
        table = Table(data, colWidths=[doc.width * 0.58, doc.width * 0.42], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), colors.HexColor("#1e40af")),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("ALIGN",         (0,0), (-1,-1), "CENTER"),
            ("FONTNAME",      (0,0), (-1,0), "DejaVuSans-Bold"),
            ("FONTSIZE",      (0,0), (-1,0), 9),
            ("BACKGROUND",    (0,1), (-1,-1), colors.white),
            ("GRID",          (0,0), (-1,-1), 0.7, colors.HexColor("#cbd5e1")),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("FONTNAME",      (0,1), (-1,-1), "DejaVuSans"),
            ("BOX",           (0,0), (-1,-1), 0.9, colors.HexColor("#93c5fd")),
            ("LEFTPADDING",   (0,0), (-1,-1), 6),
            ("RIGHTPADDING",  (0,0), (-1,-1), 6),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.16*inch))
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, mimetype="application/pdf", as_attachment=True,
                     download_name=f"procesen_cekor_{kamin.replace(' ', '_')}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.pdf")
