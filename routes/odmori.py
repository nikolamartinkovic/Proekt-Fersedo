import io
import json
import sqlite3
from collections import defaultdict
from datetime import datetime, timedelta

import openpyxl
from flask import (
    Blueprint, flash, redirect, render_template,
    request, send_file, url_for
)
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.audit import log_audit_event
from utils.db import get_db
from utils.decorators import login_required, module_required
from utils.odmori_helpers import (
    _get_odmori_for_date,
    _get_odmori_for_range,
    calc_working_days,
    get_email_log,
    ensure_manager_emails_table,
    ensure_odmor_salda_table,
    get_saldo_all,
)
from utils.odmori_notifications import (
    TIP_COLORS,
    TIP_LABELS,
    isprati_dnevni_izvestaj_otsustva,
    isprati_nedelen_izvestaj_otsustva,
)

odmori_bp = Blueprint("odmori", __name__, url_prefix="/odmori")

# ─────────────────────────────────────────────────────────────
# EMAIL КОНФИГУРАЦИЈА
# ─────────────────────────────────────────────────────────────




# ─────────────────────────────────────────────────────────────
# ПОМОШНА ФУНКЦИЈА: HTML блок за одмори
# ─────────────────────────────────────────────────────────────

@odmori_bp.route("/manager_emails", methods=["GET", "POST"])
@login_required
@module_required("odmori_manager_emails")
def odmori_manager_emails():
    conn   = get_db()
    cursor = conn.cursor()
    ensure_manager_emails_table(cursor)
    conn.commit()

    if request.method == "POST":
        action = request.form.get("action")

        if action == "add":
            email = request.form.get("email", "").strip()
            ime   = request.form.get("ime", "").strip()
            if email:
                try:
                    cursor.execute(
                        "INSERT INTO otsustva_manager_emails (email, ime, aktiven) VALUES (?,?,1)",
                        (email, ime)
                    )
                    conn.commit()
                    log_audit_event(
                        "odmori",
                        "manager_email_add",
                        status="success",
                        details=f"Додаден примач: {ime or '-'} <{email}>",
                    )
                    flash(f"Менаџерот {ime} ({email}) е успешно додаден!", "success")
                except sqlite3.IntegrityError:
                    log_audit_event(
                        "odmori",
                        "manager_email_add",
                        status="warning",
                        details=f"Дупликат email: {email}",
                    )
                    flash("Оваа email адреса веќе постои!", "danger")
                except Exception as e:
                    log_audit_event(
                        "odmori",
                        "manager_email_add",
                        status="error",
                        details=f"Грешка при додавање на {email}: {e}",
                    )
                    flash(f"Грешка: {e}", "danger")

        elif action == "toggle":
            mid = request.form.get("manager_id")
            if mid:
                try:
                    cursor.execute(
                        "UPDATE otsustva_manager_emails SET aktiven = CASE WHEN aktiven=1 THEN 0 ELSE 1 END WHERE id=?",
                        (mid,)
                    )
                    conn.commit()
                    updated_manager = cursor.execute(
                        "SELECT ime, email, aktiven FROM otsustva_manager_emails WHERE id=?",
                        (mid,),
                    ).fetchone()
                    if updated_manager:
                        status_label = "активен" if updated_manager["aktiven"] else "неактивен"
                        log_audit_event(
                            "odmori",
                            "manager_email_toggle",
                            status="success",
                            details=(
                                f"Променет статус за {updated_manager['ime'] or '-'} "
                                f"<{updated_manager['email']}> во {status_label}"
                            ),
                        )
                    flash("Статусот е ажуриран!", "success")
                except Exception as e:
                    flash(f"Грешка: {e}", "danger")

        elif action == "delete":
            mid = request.form.get("manager_id")
            if mid:
                try:
                    manager = cursor.execute(
                        "SELECT ime, email FROM otsustva_manager_emails WHERE id=?",
                        (mid,),
                    ).fetchone()
                    cursor.execute("DELETE FROM otsustva_manager_emails WHERE id=?", (mid,))
                    conn.commit()
                    details = f"Избришан примач ID {mid}"
                    if manager:
                        details = f"Избришан примач: {manager['ime'] or '-'} <{manager['email']}>"
                    log_audit_event(
                        "odmori",
                        "manager_email_delete",
                        status="success",
                        details=details,
                    )
                    flash("Менаџерот е избришан!", "success")
                except Exception as e:
                    flash(f"Грешка: {e}", "danger")

        elif action == "test_dnevni":
            conn.close()
            result = isprati_dnevni_izvestaj_otsustva() or {"success": False, "message": "\u041d\u0435\u043f\u043e\u0437\u043d\u0430\u0442\u0430 \u0433\u0440\u0435\u0448\u043a\u0430 \u043f\u0440\u0438 \u0434\u043d\u0435\u0432\u0435\u043d \u0438\u0437\u0432\u0435\u0448\u0442\u0430\u0458."}
            log_audit_event(
                "odmori",
                "manager_email_test_dneven",
                status="success" if result.get("success") else "error",
                details=result.get("message", ""),
            )
            flash(result["message"], "success" if result.get("success") else "danger")
            return redirect(url_for("odmori.odmori_manager_emails"))

        elif action == "test_nedelen":
            conn.close()
            result = isprati_nedelen_izvestaj_otsustva() or {"success": False, "message": "\u041d\u0435\u043f\u043e\u0437\u043d\u0430\u0442\u0430 \u0433\u0440\u0435\u0448\u043a\u0430 \u043f\u0440\u0438 \u043d\u0435\u0434\u0435\u043b\u0435\u043d \u0438\u0437\u0432\u0435\u0448\u0442\u0430\u0458."}
            log_audit_event(
                "odmori",
                "manager_email_test_nedelen",
                status="success" if result.get("success") else "error",
                details=result.get("message", ""),
            )
            flash(result["message"], "success" if result.get("success") else "danger")
            return redirect(url_for("odmori.odmori_manager_emails"))

    managers = cursor.execute(
        "SELECT * FROM otsustva_manager_emails ORDER BY ime, email"
    ).fetchall()
    conn.close()
    email_log = get_email_log()
    return render_template("manager_emails.html", managers=managers, email_log=email_log)


# ─────────────────────────────────────────────────────────────
# ВРАБОТЕНИ
# ─────────────────────────────────────────────────────────────

@odmori_bp.route("/vraboteni", methods=["GET", "POST"])
@login_required
@module_required("odmori_vraboteni")
def odmori_vraboteni():
    conn   = get_db()
    cursor = conn.cursor()
    godina = datetime.now().year

    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            try:
                cursor.execute("""
                    INSERT INTO vraboteni
                        (ime, prezime, maticen_broj, email, pozicija, datum_vrabotuvanje, oddel, prekin_staz)
                    VALUES (?,?,?,?,?,?,?,?)
                """, (
                    request.form.get("ime", "").strip(),
                    request.form.get("prezime", "").strip(),
                    request.form.get("maticen_broj", "").strip(),
                    request.form.get("email", "").strip(),
                    request.form.get("pozicija", "").strip(),
                    request.form.get("datum_vrabotuvanje"),
                    request.form.get("oddel", "").strip(),
                    1 if request.form.get("prekin_staz") else 0,
                ))
                conn.commit()
                flash("Вработениот е успешно додаден!", "success")
            except sqlite3.IntegrityError:
                flash("Матичниот број веќе постои!", "danger")
            except Exception as e:
                flash(f"Грешка: {e}", "danger")

        elif action == "delete":
            vraboten_id = request.form.get("vraboten_id")
            if vraboten_id:
                try:
                    cursor.execute("DELETE FROM vraboteni WHERE id=?", (vraboten_id,))
                    conn.commit()
                    flash("Вработениот е успешно избришан!", "success")
                except Exception as e:
                    flash(f"Грешка: {e}", "danger")

        elif action == "edit":
            vraboten_id = request.form.get("vraboten_id")
            if vraboten_id:
                try:
                    cursor.execute("""
                        UPDATE vraboteni
                        SET ime=?, prezime=?, maticen_broj=?, email=?,
                            pozicija=?, datum_vrabotuvanje=?, oddel=?, prekin_staz=?
                        WHERE id=?
                    """, (
                        request.form.get("ime_edit", "").strip(),
                        request.form.get("prezime_edit", "").strip(),
                        request.form.get("maticen_broj_edit", "").strip(),
                        request.form.get("email_edit", "").strip(),
                        request.form.get("pozicija_edit", "").strip(),
                        request.form.get("datum_vrabotuvanje_edit"),
                        request.form.get("oddel_edit", "").strip(),
                        1 if request.form.get("prekin_staz_edit") else 0,
                        vraboten_id,
                    ))
                    conn.commit()
                    flash("Вработениот е успешно ажуриран!", "success")
                except sqlite3.IntegrityError:
                    flash("Матичниот број веќе постои!", "danger")
                except Exception as e:
                    flash(f"Грешка: {e}", "danger")

        elif action == "edit_saldo":
            vraboten_id = request.form.get("vraboten_id")
            vkupno      = request.form.get("vkupno_dena", 20)
            godina_form = request.form.get("godina_saldo", godina)
            if vraboten_id:
                try:
                    ensure_odmor_salda_table(cursor)
                    cursor.execute("""
                        INSERT INTO odmor_salda (vraboten_id, godina, vkupno_dena)
                        VALUES (?, ?, ?)
                        ON CONFLICT(vraboten_id, godina) DO UPDATE SET vkupno_dena = excluded.vkupno_dena
                    """, (vraboten_id, int(godina_form), int(vkupno)))
                    conn.commit()
                    flash("Салдото е успешно ажурирано!", "success")
                except Exception as e:
                    flash(f"Грешка при ажурирање на салдо: {e}", "danger")

    ensure_odmor_salda_table(cursor)
    vraboteni = cursor.execute("SELECT * FROM vraboteni ORDER BY prezime, ime").fetchall()
    saldo_map = get_saldo_all(cursor, godina)
    conn.commit()
    conn.close()
    return render_template("vraboteni.html", vraboteni=vraboteni, saldo_map=saldo_map, godina=godina)


# ─────────────────────────────────────────────────────────────
# ИМПОРТ / ТЕМПЛАТ
# ─────────────────────────────────────────────────────────────

@odmori_bp.route("/vraboteni/import", methods=["POST"])
@login_required
@module_required("odmori_vraboteni")
def odmori_import_vraboteni():
    file = request.files.get("excel_file")
    if not file or not file.filename.endswith((".xlsx", ".xls")):
        flash("Ве молиме прикачете валидна Excel датотека (.xlsx)!", "danger")
        return redirect(url_for("odmori.odmori_vraboteni"))
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        expected = ["Ime *","Prezime *","Maticen broj *","Email",
                    "Pozicija","Datum vrabotuvanje (YYYY-MM-DD)","Oddel","Prekin staz (da/ne)"]
        if [str(ws.cell(1,c).value or "").strip() for c in range(1,9)] != expected:
            flash("Грешка: Хедерите не се совпаѓаат. Користете го официјалниот темплат!", "danger")
            return redirect(url_for("odmori.odmori_vraboteni"))
        conn = get_db(); cursor = conn.cursor()
        inserted = skipped = 0; errors = []
        for row_idx in range(2, ws.max_row + 1):
            vals = [ws.cell(row_idx, c).value for c in range(1, 9)]
            if all(v is None or str(v).strip() == "" for v in vals):
                continue
            ime = str(vals[0] or "").strip(); prezime = str(vals[1] or "").strip()
            maticen = str(vals[2] or "").strip()
            if not ime or not prezime or not maticen:
                errors.append(f"Ред {row_idx}: Недостасуваат задолжителни полиња."); skipped += 1; continue
            datum = str(vals[5] or "").strip() if vals[5] else None
            if datum and datum != "None":
                try:
                    if hasattr(vals[5], "strftime"): datum = vals[5].strftime("%Y-%m-%d")
                except Exception: pass
            else: datum = None
            prekin = 1 if str(vals[7] or "").strip().lower() in ("da","да","yes","1","true") else 0
            try:
                cursor.execute("""INSERT INTO vraboteni
                    (ime,prezime,maticen_broj,email,pozicija,datum_vrabotuvanje,oddel,prekin_staz)
                    VALUES (?,?,?,?,?,?,?,?)""",
                    (ime, prezime, maticen,
                     str(vals[3] or "").strip() or None,
                     str(vals[4] or "").strip() or None,
                     datum,
                     str(vals[6] or "").strip() or None,
                     prekin))
                inserted += 1
            except Exception as db_err:
                skipped += 1
                errors.append(f"Ред {row_idx} ({ime} {prezime}): {db_err}")
        conn.commit(); conn.close()
        if inserted: flash(f"Успешно увезени {inserted} вработени!", "success")
        if skipped:  flash(f"Прескокнати {skipped} редови поради грешки.", "warning")
        for err in errors[:5]: flash(err, "danger")
    except Exception as e:
        flash(f"Грешка при читање на Excel: {e}", "danger")
    return redirect(url_for("odmori.odmori_vraboteni"))


@odmori_bp.route("/vraboteni/template")
@login_required
@module_required("odmori_vraboteni")
def odmori_download_template():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Вработени"
    display = ["Ime *","Prezime *","Maticen broj *","Email","Pozicija",
               "Datum vrabotuvanje (YYYY-MM-DD)","Oddel","Prekin staz (da/ne)"]
    hfill = PatternFill("solid", start_color="B91C1C")
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    ctr   = Alignment(horizontal="center", vertical="center")
    thin  = Side(style="thin", color="CCCCCC")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, d in enumerate(display, 1):
        c = ws.cell(1, ci, d); c.font = hfont; c.fill = hfill; c.alignment = ctr; c.border = brd
    rfill = PatternFill("solid", start_color="FEF2F2")
    for ri, row in enumerate([
        ["Ana","Jovanoska","1234567890123","ana@primer.mk","Менаџер","2020-01-15","HR","ne"],
        ["Marko","Petrov","9876543210123","marko@primer.mk","Програмер","2021-06-01","IT","da"],
    ], 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val); c.font = Font(name="Arial", size=10)
            c.fill = rfill; c.border = brd; c.alignment = Alignment(vertical="center")
    for i, w in enumerate([14,16,18,26,20,32,14,22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws2 = wb.create_sheet("Упатство")
    for r, (text, bold) in enumerate([
        ("УПАТСТВО ЗА ПОПОЛНУВАЊЕ",True),("",False),
        ("Полиња означени со * се задолжителни.",False),
        ("Датум: внесете во формат YYYY-MM-DD (пр. 2024-03-15)",False),
        ("Прекин во стаж: внесете 'da' или 'ne'",False),
        ("Матичниот број мора да биде единствен.",False),
        ("Не менувајте ги имињата на колоните!",False),
    ], 1):
        c = ws2.cell(r, 1, text)
        c.font = Font(bold=bold, name="Arial", size=11 if bold else 10,
                      color="B91C1C" if bold else "000000")
    ws2.column_dimensions["A"].width = 60
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="template_vraboteni.xlsx")


# ─────────────────────────────────────────────────────────────
# КАЛЕНДАР
# ─────────────────────────────────────────────────────────────

@odmori_bp.route("/kalendar", methods=["GET", "POST"])
@login_required
@module_required("odmori_kalendar")
def odmori_kalendar():
    conn   = get_db(); cursor = conn.cursor()
    vraboteni = cursor.execute("SELECT id, ime, prezime FROM vraboteni ORDER BY prezime, ime").fetchall()
    selected_vraboten_id = request.args.get("vraboten_id", "all")
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add_neraboten":
            datum = request.form.get("datum"); ime = request.form.get("ime","").strip()
            if datum and ime:
                try:
                    cursor.execute("INSERT INTO nerabotni_deni (datum, ime) VALUES (?,?)", (datum, ime))
                    conn.commit(); flash("Неработниот ден е успешно додаден!", "success")
                except sqlite3.IntegrityError: flash("Овој датум веќе постои!", "danger")
                except Exception as e: flash(f"Грешка: {e}", "danger")
        elif action == "delete_neraboten":
            nid = request.form.get("neraboten_id")
            if nid:
                try:
                    cursor.execute("DELETE FROM nerabotni_deni WHERE id=?", (nid,))
                    conn.commit(); flash("Неработниот ден е успешно избришан!", "success")
                except Exception as e: flash(f"Грешка: {e}", "danger")
        conn.close()
        return redirect(url_for("odmori.odmori_kalendar", vraboten_id=selected_vraboten_id))

    query = """SELECT b.datum_od,b.datum_do,b.zabeleska,v.ime,v.prezime
               FROM baranja_odmor b JOIN vraboteni v ON b.vraboten_id=v.id
               WHERE b.status='approved'"""
    params = []
    if selected_vraboten_id != "all":
        query += " AND b.vraboten_id=?"; params.append(selected_vraboten_id)
    odmor_baranja = cursor.execute(query, params).fetchall()
    nerabotni     = cursor.execute("SELECT id, datum, ime FROM nerabotni_deni ORDER BY datum").fetchall()
    events = []
    for b in odmor_baranja:
        try:
            start = datetime.strptime(b["datum_od"], "%Y-%m-%d").date()
            end   = datetime.strptime(b["datum_do"], "%Y-%m-%d").date() + timedelta(days=1)
        except Exception: continue
        events.append({"title":f"{b['ime']} {b['prezime']}","start":b["datum_od"],
            "end":end.strftime("%Y-%m-%d"),"backgroundColor":"#28a745","borderColor":"#1e7e34",
            "textColor":"white","extendedProps":{"ime":b["ime"],"prezime":b["prezime"],
            "zabeleska":b["zabeleska"] or "Нема забелешка",
            "period":f"Од {start.strftime('%d-%m-%Y')} до {(end-timedelta(days=1)).strftime('%d-%m-%Y')}"}})
    for n in nerabotni:
        events.append({"title":n["ime"],"start":n["datum"],"allDay":True,
            "backgroundColor":"#dc3545","borderColor":"#c82333","textColor":"white",
            "extendedProps":{"type":"neraboten","neraboten_id":n["id"],"ime":n["ime"]}})
    current_year = datetime.now().year
    mk_months = {"January":"Јануари","February":"Февруари","March":"Март","April":"Април",
        "May":"Мај","June":"Јуни","July":"Јули","August":"Август",
        "September":"Септември","October":"Октомври","November":"Ноември","December":"Декември"}
    nerabotni_by_month = defaultdict(list)
    for n in nerabotni:
        try:
            dt = datetime.strptime(n["datum"], "%Y-%m-%d")
            if dt.year == current_year:
                nerabotni_by_month[mk_months[dt.strftime("%B")]].append(
                    {"id":n["id"],"datum":dt.strftime("%d-%m-%Y"),"ime":n["ime"]})
        except Exception: continue
    month_order = ["Јануари","Февруари","Март","Април","Мај","Јуни",
                   "Јули","Август","Септември","Октомври","Ноември","Декември"]
    nerabotni_grouped = [{"month":m,"count":len(nerabotni_by_month.get(m,[])),
        "days":sorted(nerabotni_by_month.get(m,[]),key=lambda x:x["datum"])} for m in month_order]
    conn.close()
    return render_template("kalendar.html", events_json=json.dumps(events), vraboteni=vraboteni,
        selected_vraboten_id=selected_vraboten_id, nerabotni_grouped=nerabotni_grouped,
        current_year=current_year)


# ─────────────────────────────────────────────────────────────
# ПРЕГЛЕД НА ОДМОРИ
# ─────────────────────────────────────────────────────────────

@odmori_bp.route("/pregled_odmori", methods=["GET", "POST"])
@login_required
@module_required("odmori_pregled_odmori")
def odmori_pregled_odmori():
    conn = get_db(); cursor = conn.cursor()
    godina = datetime.now().year
    if request.method == "POST":
        action = request.form.get("action"); baranje_id = request.form.get("baranje_id")
        if baranje_id and action in ("approve", "reject"):
            new_status = "approved" if action == "approve" else "rejected"
            try:
                cursor.execute("UPDATE baranja_odmor SET status=? WHERE id=?", (new_status, baranje_id))
                conn.commit(); flash(f"Барањето е {new_status.upper()}!", "success")
                if new_status == "approved":
                    try:
                        from routes.main import _isprati_odobruvanje_email
                        b = cursor.execute("""SELECT b.datum_od,b.datum_do,b.zabeleska,b.vraboten_id,
                            v.ime,v.prezime,v.email FROM baranja_odmor b
                            JOIN vraboteni v ON b.vraboten_id=v.id WHERE b.id=?""", (baranje_id,)).fetchone()
                        if b and b["email"]:
                            praznici_set = {r["datum"] for r in cursor.execute("SELECT datum FROM nerabotni_deni").fetchall()}
                            working_days = 0
                            try:
                                s = datetime.strptime(b["datum_od"],"%Y-%m-%d").date()
                                e = datetime.strptime(b["datum_do"],"%Y-%m-%d").date()
                                c = s
                                while c <= e:
                                    if c.weekday() < 5 and c.strftime("%Y-%m-%d") not in praznici_set: working_days += 1
                                    c += timedelta(days=1)
                                godina_b = s.year
                            except Exception: godina_b = datetime.now().year
                            saldo_r = cursor.execute("SELECT vkupno_dena FROM odmor_salda WHERE vraboten_id=? AND godina=?",
                                (b["vraboten_id"],godina_b)).fetchone()
                            vkupno = saldo_r["vkupno_dena"] if saldo_r else 20
                            saldo_map_now = get_saldo_all(cursor, godina_b)
                            preostanati = saldo_map_now.get(b["vraboten_id"],{}).get("preostanati",0)
                            _isprati_odobruvanje_email(vraboten_email=b["email"].strip(),
                                ime_prezime=f"{b['ime']} {b['prezime']}",datum_od=b["datum_od"],
                                datum_do=b["datum_do"],working_days=working_days,zabeleska=b["zabeleska"] or "",
                                odobren_od="Менаџмент",vkupno_dena=vkupno,preostanati=preostanati,godina=godina_b)
                    except Exception as mail_err: print(f"[EMAIL ODOBR] Грешка: {mail_err}")
            except Exception as e: flash(f"Грешка: {e}", "danger")
        elif action == "edit" and baranje_id:
            datum_od = request.form.get("datum_od"); datum_do = request.form.get("datum_do")
            zabeleska = request.form.get("zabeleska","").strip()
            if datum_od and datum_do:
                try:
                    od = datetime.strptime(datum_od,"%Y-%m-%d").date()
                    do = datetime.strptime(datum_do,"%Y-%m-%d").date()
                    if od > do: flash('Датумот "Од" не може да биде после "До"!', "danger")
                    else:
                        cursor.execute("UPDATE baranja_odmor SET datum_od=?,datum_do=?,zabeleska=? WHERE id=?",
                            (datum_od,datum_do,zabeleska,baranje_id))
                        conn.commit(); flash("Барањето е успешно изменето!", "success")
                except Exception as e: flash(f"Грешка: {e}", "danger")
        elif action == "cancel" and baranje_id:
            try:
                cursor.execute("UPDATE baranja_odmor SET status='otkazano' WHERE id=?", (baranje_id,))
                conn.commit(); flash("Барањето е успешно откажано!", "success")
            except Exception as e: flash(f"Грешка: {e}", "danger")

    praznici = {r["datum"] for r in cursor.execute("SELECT datum FROM nerabotni_deni").fetchall()}
    baranja_raw = cursor.execute("""SELECT b.id,b.vraboten_id,b.datum_od,b.datum_do,b.status,b.zabeleska,
        b.podneseno_od,b.podneseno_na,v.ime,v.prezime FROM baranja_odmor b
        JOIN vraboteni v ON b.vraboten_id=v.id ORDER BY b.podneseno_na DESC""").fetchall()
    baranja = []
    for b in baranja_raw:
        try: start = datetime.strptime(b["datum_od"],"%Y-%m-%d").date(); end = datetime.strptime(b["datum_do"],"%Y-%m-%d").date()
        except Exception: start = end = None
        total_days = working_days = 0
        if start and end and start <= end:
            total_days = (end - start).days + 1
            cur = start
            while cur <= end:
                if cur.weekday() < 5 and cur.strftime("%Y-%m-%d") not in praznici: working_days += 1
                cur += timedelta(days=1)
        baranja.append({"id":b["id"],"vraboten_id":b["vraboten_id"],"ime":b["ime"],"prezime":b["prezime"],
            "datum_od_formatted":start.strftime("%d-%m-%Y") if start else "—",
            "datum_do_formatted":end.strftime("%d-%m-%Y") if end else "—",
            "datum_od_year":start.year if start else godina,"status":b["status"],
            "zabeleska":b["zabeleska"] or "—","podneseno_od":b["podneseno_od"] or "—",
            "podneseno_na":b["podneseno_na"] or "—","total_days":total_days,"working_days":working_days})
    saldo_map = get_saldo_all(cursor, godina)
    conn.commit(); conn.close()
    return render_template("pregled_odmori.html", baranja=baranja, saldo_map=saldo_map, godina=godina)


# ─────────────────────────────────────────────────────────────
# СЕКОЈДНЕВНИ ОТСУСТВА
# ─────────────────────────────────────────────────────────────

@odmori_bp.route("/sekojdnevni_otsustva", methods=["GET", "POST"])
@login_required
@module_required("odmori_sekojdnevni_otsustva")
def odmori_sekojdnevni_otsustva():
    conn = get_db(); cursor = conn.cursor()
    vraboteni = cursor.execute("SELECT id, ime, prezime FROM vraboteni ORDER BY prezime, ime").fetchall()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            vraboten_id = request.form.get("vraboten_id"); datum = request.form.get("datum")
            tip = request.form.get("tip"); casovi = float(request.form.get("casovi", 8.0))
            plateno = 1 if request.form.get("plateno") else 0
            zabeleska = request.form.get("zabeleska","").strip()
            if vraboten_id and datum and tip:
                try:
                    cursor.execute("""INSERT INTO sekojdnevni_otsustva
                        (vraboten_id,datum,tip,casovi,plateno,zabeleska) VALUES (?,?,?,?,?,?)""",
                        (vraboten_id,datum,tip,casovi,plateno,zabeleska))
                    conn.commit(); flash("Отсуството е успешно додадено!", "success")
                except Exception as e: flash(f"Грешка: {e}", "danger")
        elif action == "delete":
            oid = request.form.get("otsustvo_id")
            if oid:
                try:
                    cursor.execute("DELETE FROM sekojdnevni_otsustva WHERE id=?", (oid,))
                    conn.commit(); flash("Отсуството е успешно избришано!", "success")
                except Exception as e: flash(f"Грешка: {e}", "danger")
        elif action == "edit":
            oid = request.form.get("otsustvo_id")
            if oid:
                try:
                    cursor.execute("""UPDATE sekojdnevni_otsustva
                        SET vraboten_id=?,datum=?,tip=?,casovi=?,plateno=?,zabeleska=? WHERE id=?""",
                        (request.form.get("vraboten_id_edit"),request.form.get("datum_edit"),
                         request.form.get("tip_edit"),float(request.form.get("casovi_edit",8.0)),
                         1 if request.form.get("plateno_edit") else 0,
                         request.form.get("zabeleska_edit","").strip(),oid))
                    conn.commit(); flash("Отсуството е успешно ажурирано!", "success")
                except Exception as e: flash(f"Грешка: {e}", "danger")

    otsustva = cursor.execute("""SELECT o.id,o.datum,o.tip,o.casovi,o.plateno,o.zabeleska,
        v.ime,v.prezime,o.vraboten_id FROM sekojdnevni_otsustva o
        JOIN vraboteni v ON o.vraboten_id=v.id ORDER BY o.datum DESC""").fetchall()
    today = datetime.now().date(); today_str = today.strftime("%Y-%m-%d")
    week_start = today - timedelta(days=today.weekday()); week_end = week_start + timedelta(days=6)
    ws_str = week_start.strftime("%Y-%m-%d"); we_str = week_end.strftime("%Y-%m-%d")
    izvestaj = cursor.execute("""SELECT o.id,v.ime,v.prezime,o.datum,o.tip,o.zabeleska,o.casovi
        FROM sekojdnevni_otsustva o JOIN vraboteni v ON o.vraboten_id=v.id
        WHERE o.datum=? ORDER BY v.prezime,v.ime,o.tip""",(today_str,)).fetchall()
    nedelen_raw = cursor.execute("""SELECT v.ime,v.prezime,o.datum,o.tip,o.casovi,o.plateno,
        o.zabeleska,o.vraboten_id FROM sekojdnevni_otsustva o
        JOIN vraboteni v ON o.vraboten_id=v.id
        WHERE o.datum BETWEEN ? AND ? ORDER BY o.datum ASC,v.prezime,v.ime""",(ws_str,we_str)).fetchall()
    mk_days = {0:"Пон",1:"Вто",2:"Сре",3:"Чет",4:"Пет",5:"Саб",6:"Нед"}
    nedelen_po_vraboten = defaultdict(lambda:{"ime":"","prezime":"","dena":[],"vkupno_casovi":0.0})
    for r in nedelen_raw:
        vid = r["vraboten_id"]
        nedelen_po_vraboten[vid]["ime"] = r["ime"]; nedelen_po_vraboten[vid]["prezime"] = r["prezime"]
        nedelen_po_vraboten[vid]["vkupno_casovi"] += r["casovi"]
        try: d = datetime.strptime(r["datum"],"%Y-%m-%d").date(); day_label = f"{mk_days[d.weekday()]} {d.strftime('%d-%m')}"
        except Exception: day_label = r["datum"]
        nedelen_po_vraboten[vid]["dena"].append({"dan":day_label,"tip":r["tip"],"casovi":r["casovi"],"plateno":r["plateno"]})
    tip_summary = defaultdict(lambda:{"count":0,"casovi":0.0})
    for r in nedelen_raw: tip_summary[r["tip"]]["count"] += 1; tip_summary[r["tip"]]["casovi"] += r["casovi"]
    week_days_list = [{"label":f"{mk_days[i]} {(week_start+timedelta(days=i)).strftime('%d-%m')}",
        "datum":(week_start+timedelta(days=i)).strftime("%Y-%m-%d"),
        "is_today":(week_start+timedelta(days=i))==today,"is_weekend":i>=5} for i in range(7)]

    praznici      = {r["datum"] for r in cursor.execute("SELECT datum FROM nerabotni_deni").fetchall()}
    odmori_denes  = _get_odmori_for_date(cursor, today_str, praznici)
    odmori_nedela = _get_odmori_for_range(cursor, ws_str, we_str, praznici)

    conn.close()
    return render_template("sekojdnevni_otsustva.html", otsustva=otsustva, vraboteni=vraboteni,
        izvestaj=izvestaj, nedelen_raw=list(nedelen_raw), nedelen_po_vraboten=dict(nedelen_po_vraboten),
        tip_summary=dict(tip_summary), week_days_list=week_days_list,
        week_start=week_start.strftime("%d-%m-%Y"), week_end=week_end.strftime("%d-%m-%Y"),
        today=today.strftime("%d-%m-%Y"), today_str=today_str,
        odmori_denes=odmori_denes, odmori_nedela=odmori_nedela)
    
@odmori_bp.route("/nedeli")
@login_required
@module_required("odmori_sekojdnevni_otsustva")
def odmori_nedeli():
    conn   = get_db()
    cursor = conn.cursor()

    # Земи ги сите отсуства групирани по недела
    otsustva = cursor.execute("""
        SELECT o.id, o.datum, o.tip, o.casovi, o.plateno, o.zabeleska,
               v.ime, v.prezime, o.vraboten_id,
               strftime('%Y', o.datum) as godina,
               strftime('%W', o.datum) as kn
        FROM sekojdnevni_otsustva o
        JOIN vraboteni v ON o.vraboten_id = v.id
        ORDER BY o.datum DESC
    """).fetchall()

    praznici     = {r["datum"] for r in cursor.execute("SELECT datum FROM nerabotni_deni").fetchall()}
    vraboteni    = cursor.execute("SELECT id, ime, prezime FROM vraboteni ORDER BY prezime, ime").fetchall()

    # Земи ги сите одмори
    odmori_all = cursor.execute("""
        SELECT v.ime, v.prezime, b.datum_od, b.datum_do, b.zabeleska, b.vraboten_id
        FROM baranja_odmor b
        JOIN vraboteni v ON b.vraboten_id = v.id
        WHERE b.status = 'approved'
        ORDER BY b.datum_od DESC
    """).fetchall()

    conn.close()

    # Групирај по недела
    from collections import defaultdict
    import datetime as dt

    nedeli = defaultdict(lambda: {
        "otsustva": [],
        "odmori":   [],
        "datum_od": None,
        "datum_do": None,
        "kn":       None,
        "godina":   None,
    })

    for o in otsustva:
        try:
            d         = dt.datetime.strptime(o["datum"], "%Y-%m-%d").date()
            week_start = d - dt.timedelta(days=d.weekday())
            week_end   = week_start + dt.timedelta(days=4)
            kn         = d.isocalendar()[1]
            godina     = d.year
            key        = f"{godina}-{kn:02d}"

            nedeli[key]["kn"]       = kn
            nedeli[key]["godina"]   = godina
            nedeli[key]["datum_od"] = week_start.strftime("%d-%m-%Y")
            nedeli[key]["datum_do"] = week_end.strftime("%d-%m-%Y")
            nedeli[key]["otsustva"].append(dict(o))
        except Exception:
            continue

    # Додај одмори во соодветните недели
    for o in odmori_all:
        try:
            start = dt.datetime.strptime(o["datum_od"], "%Y-%m-%d").date()
            end   = dt.datetime.strptime(o["datum_do"], "%Y-%m-%d").date()

            # Прошетај низ неделите кои ги покрива одморот
            cur = start - dt.timedelta(days=start.weekday())
            while cur <= end:
                kn     = cur.isocalendar()[1]
                godina = cur.year
                key    = f"{godina}-{kn:02d}"

                week_end_d = cur + dt.timedelta(days=4)
                if nedeli[key]["datum_od"] is None:
                    nedeli[key]["kn"]       = kn
                    nedeli[key]["godina"]   = godina
                    nedeli[key]["datum_od"] = cur.strftime("%d-%m-%Y")
                    nedeli[key]["datum_do"] = week_end_d.strftime("%d-%m-%Y")

                wd = calc_working_days(o["datum_od"], o["datum_do"], praznici)
                odmor_entry = {
                    "ime":          o["ime"],
                    "prezime":      o["prezime"],
                    "datum_od":     o["datum_od"],
                    "datum_do":     o["datum_do"],
                    "zabeleska":    o["zabeleska"] or "",
                    "working_days": wd,
                }
                # Не дуплирај
                if odmor_entry not in nedeli[key]["odmori"]:
                    nedeli[key]["odmori"].append(odmor_entry)

                cur += dt.timedelta(days=7)
        except Exception:
            continue

    # Сортирај по клуч опаѓачки
    nedeli_sorted = sorted(nedeli.items(), key=lambda x: x[0], reverse=True)

    selected_kn = request.args.get("kn", None)

    return render_template(
        "odmori_nedeli.html",
        nedeli=nedeli_sorted,
        selected_kn=selected_kn,
        TIP_COLORS=TIP_COLORS,
        TIP_LABELS=TIP_LABELS,
    )
