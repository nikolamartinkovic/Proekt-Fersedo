import io
import sqlite3
from datetime import datetime, date, timedelta
from collections import defaultdict
from flask import Blueprint, render_template, request, flash, redirect, url_for, jsonify, send_file, session
import openpyxl
from utils.db import get_db
from utils.decorators import admin_or_module_required, login_required
from utils.zalihi_profakturi import (
    create_profaktura,
    get_izvoz_artikli,
    get_pending_profakturi,
    get_pregled_data,
    get_profakturi_nova_artikli,
    approve_profaktura,
    reject_profaktura,
)
from utils.zalihi_workflow import (
    approve_all_pending_exports,
    approve_pending_export,
    execute_storno,
    get_pending_exports,
    get_storno_history,
    reject_pending_export,
    submit_pending_exports,
)
from utils.zalihi_exports import (
    build_dodadeni_po_nedeli_workbook,
    build_istorija_workbook,
    build_izvozi_po_nedeli_view_model,
    build_izvozi_po_nedeli_workbook,
    group_izvozi_po_nedeli,
    prepare_dodadeni_po_nedeli_export_rows,
    prepare_istorija_rows,
)

zalihi_bp = Blueprint('zalihi', __name__, url_prefix='/zalihi')

# Помошна функција за недела


def _normalize_excel_header(value):
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def _parse_dashboard_export_excel_date(value):
    if value is None:
        raise ValueError("Недостасува датум.")

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, (int, float)):
        if value <= 0:
            raise ValueError("Датумот не е валиден.")
        excel_base = datetime(1899, 12, 30)
        return (excel_base + timedelta(days=float(value))).date().isoformat()

    text = str(value).strip()
    if not text or text.lower() == "none":
        raise ValueError("Недостасува датум.")

    normalized = text.replace(".", "-").replace("/", "-")
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(normalized, fmt).date().isoformat()
        except ValueError:
            continue

    raise ValueError(f"Неподдржан датум: {text}")


@zalihi_bp.route("/get_artikal/<pn>")
@login_required
def get_artikal(pn):
    try:
        print(f"🔍 DEBUG: get_artikal called with pn='{pn}'")
        conn = get_db()
        cursor = conn.cursor()
        pn_normalized = pn.upper().strip()
        print(f"🔍 Normalized PN: '{pn_normalized}'")
        result = cursor.execute("""
            SELECT id, part_number, ime
            FROM parts
            WHERE UPPER(part_number) = ?
            LIMIT 1
        """, (pn_normalized,)).fetchone()
        print(f"🔍 Result: {result}")
        conn.close()
        if result:
            result_dict = dict(result) if hasattr(result, 'keys') else result
            print(f"✅ УСПЕШНО ПРОНАЈДЕНО: {result_dict}")
            return jsonify({
                'success': True,
                'id': result_dict['id'],
                'pn': result_dict['part_number'],
                'ime': result_dict['ime'] if result_dict['ime'] else 'Без назив'
            })
        print(f"⚠️ НЕ ПРОНАЈДЕНО")
        return jsonify({'success': False}), 404
    except Exception as e:
        print(f"❌ ГРЕШКА: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────────────────────
# ГЛАВНА СТРАНИЦА – САМО КАРТИЧКИ
# ─────────────────────────────────────────────────────────────
@zalihi_bp.route("/", methods=["GET"])
@login_required
def zalihi():
    return render_template("zalihi.html")


def _get_dashboard_export_target(year):
    conn = get_db()
    cursor = conn.cursor()
    try:
        row = cursor.execute(
            "SELECT target_total FROM dashboard_izvozi_targets WHERE year = ? LIMIT 1",
            (int(year),),
        ).fetchone()
        if row and row.get("target_total") is not None:
            try:
                return max(1, int(row["target_total"]))
            except (TypeError, ValueError):
                return 4000
        return 4000
    finally:
        conn.close()


def _build_dashboard_export_chart_payload(year=None, cursor=None):
    selected_year = int(year or date.today().year)

    close_conn = False
    if cursor is None:
        conn = get_db()
        cursor = conn.cursor()
        close_conn = True
    try:
        target_row = cursor.execute(
            "SELECT target_total FROM dashboard_izvozi_targets WHERE year = ? LIMIT 1",
            (selected_year,),
        ).fetchone()
        if target_row and target_row.get("target_total") is not None:
            try:
                target_total = max(1, int(target_row["target_total"]))
            except (TypeError, ValueError):
                target_total = 4000
        else:
            target_total = 4000

        rows = cursor.execute(
            """
            SELECT substr(datum, 1, 7) AS month_key, SUM(kolicina) AS total
            FROM dashboard_izvozi
            WHERE substr(datum, 1, 4) = ?
            GROUP BY substr(datum, 1, 7)
            ORDER BY month_key ASC
            """,
            (str(selected_year),),
        ).fetchall()
    finally:
        if close_conn:
            conn.close()

    monthly_exports = [0] * 12
    for row in rows:
        month_key = row["month_key"] if hasattr(row, "keys") else row[0]
        if not month_key or len(str(month_key)) < 7:
            continue
        try:
            month_index = int(str(month_key)[5:7]) - 1
        except (TypeError, ValueError):
            continue
        if 0 <= month_index < 12:
            monthly_exports[month_index] = int(row["total"] or 0)

    cumulative_exports = []
    running_total = 0
    for value in monthly_exports:
        running_total += int(value or 0)
        cumulative_exports.append(running_total)

    target_line = [round(target_total * ((idx + 1) / 12), 1) for idx in range(12)]
    remaining_to_target = max(target_total - running_total, 0)
    progress = round((running_total / target_total) * 100, 1) if target_total else 0

    return {
        "year": selected_year,
        "labels": ["Јан", "Фев", "Мар", "Апр", "Мај", "Јун", "Јул", "Авг", "Сеп", "Окт", "Ное", "Дек"],
        "monthly_exports": monthly_exports,
        "cumulative_exports": cumulative_exports,
        "target_line": target_line,
        "total_exported": running_total,
        "target_total": target_total,
        "remaining_to_target": remaining_to_target,
        "progress": progress,
    }


@zalihi_bp.route("/dashboard-izvozi", methods=["GET"])
@login_required
@admin_or_module_required("zalihi")
def dashboard_izvozi_legacy():
    query_args = request.args.to_dict(flat=True)
    return redirect(url_for("zalihi.dashboard_izvozi", **query_args), code=302)


@zalihi_bp.route("/izvozi", methods=["GET", "POST"])
@login_required
@admin_or_module_required("zalihi")
def dashboard_izvozi():
    conn = get_db()
    cursor = conn.cursor()

    if request.method == "POST":
        datum = (request.form.get("datum") or date.today().isoformat()).strip()
        zabeleska = (request.form.get("zabeleska") or "").strip()
        try:
            kolicina = int(request.form.get("kolicina", 0))
        except (TypeError, ValueError):
            kolicina = 0

        try:
            datetime.strptime(datum, "%Y-%m-%d")
        except ValueError:
            flash("Невалиден датум за dashboard извоз.", "danger")
        else:
            if kolicina <= 0:
                flash("Количината мора да биде поголема од 0.", "danger")
            else:
                try:
                    cursor.execute(
                        """
                        INSERT INTO dashboard_izvozi (datum, kolicina, zabeleska, username)
                        VALUES (?, ?, ?, ?)
                        """,
                        (datum, kolicina, zabeleska, session["user"]),
                    )
                    conn.commit()
                    flash("Успешно е додаден dashboard извоз.", "success")
                    conn.close()
                    return redirect(url_for("zalihi.dashboard_izvozi"))
                except Exception as exc:
                    conn.rollback()
                    flash(f"Грешка при зачувување: {str(exc)}", "danger")

    try:
        year = int(request.args.get("year") or date.today().year)
    except (TypeError, ValueError):
        year = date.today().year

    chart_payload = _build_dashboard_export_chart_payload(year, cursor=cursor)
    target_total = chart_payload["target_total"]
    total_year = chart_payload["total_exported"]

    entries = cursor.execute(
        """
        SELECT id, datum, kolicina, zabeleska, username, created_at
        FROM dashboard_izvozi
        WHERE substr(datum, 1, 4) = ?
        ORDER BY datum DESC, id DESC
        LIMIT 200
        """,
        (str(year),),
    ).fetchall()
    conn.close()

    return render_template(
        "zalihi_dashboard_izvozi.html",
        entries=entries,
        year=year,
        total_year=total_year,
        target_total=target_total,
        remaining_to_target=max(0, target_total - total_year),
        target_progress=min(100, round((total_year / target_total) * 100, 1)) if target_total else 0,
        entry_count=len(entries),
        today=date.today().isoformat(),
    )


@zalihi_bp.route("/izvozi/import", methods=["POST"])
@login_required
@admin_or_module_required("zalihi")
def dashboard_izvozi_import():
    file = request.files.get("excel_file")
    try:
        current_year = int(request.form.get("year") or date.today().year)
    except (TypeError, ValueError):
        current_year = date.today().year

    if not file or not file.filename or not file.filename.lower().endswith(".xlsx"):
        flash("Прикачи валидна .xlsx Excel датотека со колони datum и kolicina.", "danger")
        return redirect(url_for("zalihi.dashboard_izvozi", year=current_year))

    try:
        workbook = openpyxl.load_workbook(file, data_only=True)
        worksheet = workbook.active
    except Exception as exc:
        flash(f"Excel датотеката не може да се прочита: {exc}", "danger")
        return redirect(url_for("zalihi.dashboard_izvozi", year=current_year))

    headers = [_normalize_excel_header(worksheet.cell(1, col_idx).value) for col_idx in range(1, worksheet.max_column + 1)]

    datum_idx = next(
        (
            idx for idx, header in enumerate(headers)
            if header in {"datum", "датум"} or header.startswith("datum") or header.startswith("датум")
        ),
        None,
    )
    kolicina_idx = next(
        (
            idx for idx, header in enumerate(headers)
            if header in {"kolicina", "количина", "kolicinapcs", "количинаpcs", "quantity"}
            or header.startswith("kolicina")
            or header.startswith("количина")
        ),
        None,
    )

    if datum_idx is None or kolicina_idx is None:
        flash("Excel-от мора да има две колони: datum и kolicina.", "danger")
        return redirect(url_for("zalihi.dashboard_izvozi", year=current_year))

    conn = get_db()
    cursor = conn.cursor()
    imported = 0
    skipped = 0
    errors = []

    try:
        for row_idx in range(2, worksheet.max_row + 1):
            raw_datum = worksheet.cell(row_idx, datum_idx + 1).value
            raw_kolicina = worksheet.cell(row_idx, kolicina_idx + 1).value

            if (raw_datum is None or str(raw_datum).strip() == "") and (raw_kolicina is None or str(raw_kolicina).strip() == ""):
                continue

            try:
                datum = _parse_dashboard_export_excel_date(raw_datum)
                if raw_kolicina is None or str(raw_kolicina).strip() == "":
                    raise ValueError("Недостасува количина.")

                if isinstance(raw_kolicina, str):
                    normalized_qty = raw_kolicina.strip().replace(",", ".")
                    kolicina = int(float(normalized_qty))
                else:
                    kolicina = int(float(raw_kolicina))

                if kolicina <= 0:
                    raise ValueError("Количината мора да биде поголема од 0.")

                cursor.execute(
                    """
                    INSERT INTO dashboard_izvozi (datum, kolicina, zabeleska, username)
                    VALUES (?, ?, ?, ?)
                    """,
                    (datum, kolicina, "", session.get("user", "")),
                )
                imported += 1
            except Exception as row_exc:
                skipped += 1
                errors.append(f"Ред {row_idx}: {row_exc}")

        conn.commit()
    except Exception as exc:
        conn.rollback()
        flash(f"Грешка при import од Excel: {exc}", "danger")
        conn.close()
        return redirect(url_for("zalihi.dashboard_izvozi", year=current_year))
    finally:
        conn.close()

    if imported:
        flash(f"Успешно се импортирани {imported} извозни записи.", "success")
    if skipped:
        flash(f"Прескокнати се {skipped} редови поради невалидни податоци.", "warning")
        for err in errors[:5]:
            flash(err, "danger")
    if not imported and not skipped:
        flash("Во Excel датотеката нема редови за import.", "warning")

    return redirect(url_for("zalihi.dashboard_izvozi", year=current_year))


@zalihi_bp.route("/izvozi/template")
@login_required
@admin_or_module_required("zalihi")
def dashboard_izvozi_template():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Dashboard izvozi"
    worksheet.append(["datum", "kolicina"])
    worksheet.append(["2026-01-01", 111])
    worksheet.append(["2026-02-01", 160])
    worksheet.append(["2026-03-14", 50])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="dashboard_izvozi_template.xlsx",
    )


@zalihi_bp.route("/izvozi/target", methods=["POST"])
@login_required
@admin_or_module_required("zalihi")
def dashboard_izvozi_target():
    if not session.get("is_admin"):
        flash("Само администратор може да го менува target-от.", "danger")
        return redirect(url_for("zalihi.dashboard_izvozi", year=year))

    try:
        year = int(request.form.get("year") or date.today().year)
        target_total = int(request.form.get("target_total") or 4000)
    except (TypeError, ValueError):
        flash("Невалидни податоци за target.", "danger")
        return redirect(url_for("zalihi.dashboard_izvozi", year=year))

    if target_total <= 0:
        flash("Target мора да биде број поголем од 0.", "danger")
        return redirect(url_for("zalihi.dashboard_izvozi", year=year))

    conn = get_db()
    cursor = conn.cursor()
    try:
        existing = cursor.execute(
            "SELECT year FROM dashboard_izvozi_targets WHERE year = ?",
            (year,),
        ).fetchone()
        if existing:
            cursor.execute(
                """
                UPDATE dashboard_izvozi_targets
                SET target_total = ?, updated_by = ?, updated_at = CURRENT_TIMESTAMP
                WHERE year = ?
                """,
                (target_total, session.get("user", ""), year),
            )
        else:
            cursor.execute(
                """
                INSERT INTO dashboard_izvozi_targets (year, target_total, updated_by)
                VALUES (?, ?, ?)
                """,
                (year, target_total, session.get("user", "")),
            )
        conn.commit()
        flash(f"Target-ot za {year} godina e zacuvan na {target_total} pcs.", "success")
    except Exception as exc:
        conn.rollback()
        flash(f"Greška pri zacuvuvanje na target: {str(exc)}", "danger")
    finally:
        conn.close()

    return redirect(url_for("zalihi.dashboard_izvozi", year=year))


@zalihi_bp.route("/izvozi/<int:entry_id>/delete", methods=["POST"])
@login_required
@admin_or_module_required("zalihi")
def dashboard_izvozi_delete(entry_id):
    try:
        year = int(request.form.get("year") or date.today().year)
    except (TypeError, ValueError):
        year = date.today().year

    if not session.get("is_admin"):
        flash("Само администратор може да брише dashboard извоз.", "danger")
        return redirect(url_for("zalihi.dashboard_izvozi"))

    conn = get_db()
    cursor = conn.cursor()
    try:
        row = cursor.execute(
            "SELECT id FROM dashboard_izvozi WHERE id = ?",
            (entry_id,),
        ).fetchone()
        if not row:
            flash("Записот не е пронајден.", "warning")
            conn.close()
            return redirect(url_for("zalihi.dashboard_izvozi", year=year))

        cursor.execute("DELETE FROM dashboard_izvozi WHERE id = ?", (entry_id,))
        conn.commit()
        flash("Записот е избришан.", "success")
    except Exception as exc:
        conn.rollback()
        flash(f"Грешка при бришење: {str(exc)}", "danger")
    finally:
        conn.close()

    return redirect(url_for("zalihi.dashboard_izvozi", year=year))


@zalihi_bp.route("/izvozi/delete-all", methods=["POST"])
@login_required
@admin_or_module_required("zalihi")
def dashboard_izvozi_delete_all():
    try:
        year = int(request.form.get("year") or date.today().year)
    except (TypeError, ValueError):
        year = date.today().year

    if not session.get("is_admin"):
        flash("Само администратор може групно да ги брише dashboard извозите.", "danger")
        return redirect(url_for("zalihi.dashboard_izvozi", year=year))

    conn = get_db()
    cursor = conn.cursor()
    try:
        total_row = cursor.execute(
            "SELECT COUNT(*) AS total FROM dashboard_izvozi WHERE substr(datum, 1, 4) = ?",
            (str(year),),
        ).fetchone()
        total = int(total_row["total"]) if total_row and total_row["total"] is not None else 0

        if total <= 0:
            flash(f"Нема извозни записи за бришење за {year}.", "warning")
            return redirect(url_for("zalihi.dashboard_izvozi", year=year))

        cursor.execute(
            "DELETE FROM dashboard_izvozi WHERE substr(datum, 1, 4) = ?",
            (str(year),),
        )
        conn.commit()
        flash(f"Избришани се {total} извозни записи за {year}.", "success")
    except Exception as exc:
        conn.rollback()
        flash(f"Грешка при групно бришење: {str(exc)}", "danger")
    finally:
        conn.close()

    return redirect(url_for("zalihi.dashboard_izvozi", year=year))


# ─────────────────────────────────────────────────────────────
# ДОДАДИ ЗАЛИХА
# ─────────────────────────────────────────────────────────────
@zalihi_bp.route("/dodadi", methods=["GET", "POST"])
@login_required
def dodadi():
    conn = get_db()
    cursor = conn.cursor()

    if request.method == "POST":
        artikl_id = request.form.get("artikl_id")
        try:
            kolicina = int(request.form.get("kolicina", 0))
        except:
            kolicina = 0
        try:
            cena = float(request.form.get("cena", 0))
        except:
            cena = 0.0
        datum = request.form.get("datum") or datetime.now().strftime("%Y-%m-%d")
        plateno_str = request.form.get("plateno", "0")
        plateno = 1 if plateno_str == "1" else 0
        zabeleska = request.form.get("zabeleska", "").strip()

        print(f"🔍 DEBUG: artikl_id={artikl_id}, plateno={plateno}, kolicina={kolicina}")

        if not artikl_id or kolicina <= 0:
            flash("Внеси валиден артикл и количина!", "danger")
        else:
            try:
                cursor.execute("""
                    INSERT INTO zaliha_dodadi (artikl_id, kolicina, cena, datum, plateno, username, zabeleska)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (artikl_id, kolicina, cena, datum, plateno, session["user"], zabeleska))

                cursor.execute("""
    INSERT INTO zaliha_dodadi_log (datum, username, pn, ime, kolicina, tip, zabeleska)
    SELECT ?, ?, p.part_number, COALESCE(NULLIF(TRIM(p.ime), ''), 'Непознат артикл'), ?, ?, ?
    FROM parts p WHERE p.id = ?
""", (datum, session["user"], kolicina, "Додавање", zabeleska, artikl_id))

                conn.commit()
                flash("Успешно додадено во залиха!", "success")
            except Exception as e:
                print(f"❌ Грешка: {e}")
                flash(f"Грешка при додавање: {str(e)}", "danger")
                conn.rollback()

        conn.close()
        return redirect(url_for("zalihi.izvoz"))

    artikli = cursor.execute("SELECT id, part_number, ime FROM parts ORDER BY part_number").fetchall()
    conn.close()
    return render_template("zalihi_dodadi.html", artikli=artikli, today=date.today().isoformat())


# ─────────────────────────────────────────────────────────────
# ПРЕФРЛАЊЕ НА ЗАЛИХА (платена → неплатена или обратно)
# ─────────────────────────────────────────────────────────────
@zalihi_bp.route("/premesti", methods=["POST"])
@login_required
def premesti():
    conn = None
    try:
        data = request.get_json()
        print(f"\n🔍 ПРЕМЕСТИ: Приспеал JSON: {data}")

        artikl_id = int(data.get("artikl_id", 0))
        pn = data.get("pn", "").strip().upper()
        kolicina = int(data.get("kolicina", 0))
        od_plateno = int(data.get("od_plateno", 0))
        na_plateno = 1 - od_plateno
        zabeleska = data.get("zabeleska", "Префрлање од корисник").strip()

        if artikl_id <= 0 or kolicina <= 0:
            return jsonify({'success': False, 'message': 'Невалиден ID или количина'}), 400

        conn = get_db()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        row = cursor.execute("""
            SELECT COALESCE(SUM(kolicina), 0) AS total
            FROM zaliha_dodadi
            WHERE artikl_id = ? AND plateno = ?
        """, (artikl_id, od_plateno)).fetchone()

        current = row['total'] if row else 0

        if kolicina > current:
            conn.close()
            return jsonify({'success': False, 'message': f'Немате доволно залиха. Расположливо: {current}'}), 400

        remaining = kolicina
        rows = cursor.execute("""
            SELECT id, kolicina FROM zaliha_dodadi
            WHERE artikl_id = ? AND plateno = ? AND kolicina > 0
            ORDER BY datum ASC
        """, (artikl_id, od_plateno)).fetchall()

        for row_data in rows:
            if remaining <= 0:
                break
            row_id = row_data['id']
            row_kolicina = int(row_data['kolicina'])

            if row_kolicina <= remaining:
                cursor.execute("DELETE FROM zaliha_dodadi WHERE id = ?", (row_id,))
                remaining -= row_kolicina
            else:
                cursor.execute("UPDATE zaliha_dodadi SET kolicina = kolicina - ? WHERE id = ?", (remaining, row_id))
                remaining = 0

        cursor.execute("""
            INSERT INTO zaliha_dodadi (artikl_id, kolicina, cena, datum, plateno, username, zabeleska)
            VALUES (?, ?, 0, ?, ?, ?, ?)
        """, (artikl_id, kolicina, datetime.now().strftime("%Y-%m-%d"), na_plateno, session["user"], zabeleska))

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'message': f'Префрлено {kolicina} единици од {"платена" if od_plateno else "неплатена"} во {"платена" if na_plateno else "неплатена"}!'
        })

    except Exception as e:
        print(f"❌ КРИТИЧНА ГРЕШКА: {e}")
        import traceback
        traceback.print_exc()
        if conn is not None:
            try:
                conn.rollback()
                conn.close()
            except:
                pass
        return jsonify({'success': False, 'message': f'Грешка: {str(e)}'}), 500


# ─────────────────────────────────────────────────────────────
# ИЗВОЗ  →  запишува во pending (НЕ ги брише залихите веднаш)
# ─────────────────────────────────────────────────────────────
@zalihi_bp.route("/izvoz", methods=["GET", "POST"])
@login_required
def izvoz():
    if request.method == "POST":
        response, status_code = submit_pending_exports(request.get_json() or {}, session["user"])
        return jsonify(response), status_code

    try:
        artikli = get_izvoz_artikli()
    except Exception as e:
        print(f"Грешка при читање артикли за извоз: {e}")
        import traceback
        traceback.print_exc()
        artikli = []

    return render_template("zalihi_izvoz.html", artikli=artikli, today=date.today().isoformat())

@zalihi_bp.route("/pregled")
@login_required
def pregled():
    plateni, neplateni, profaktura_zaliha = get_pregled_data()
    return render_template(
        "zalihi_pregled.html",
        plateni=plateni,
        neplateni=neplateni,
        profaktura_zaliha=profaktura_zaliha,
    )

@zalihi_bp.route("/debug")
@login_required
def debug():
    conn = get_db()
    cursor = conn.cursor()

    site = cursor.execute("""
        SELECT d.id, d.artikl_id, d.kolicina, d.plateno, d.datum, d.username,
               p.part_number, p.ime
        FROM zaliha_dodadi d
        LEFT JOIN parts p ON d.artikl_id = p.id
        ORDER BY d.id DESC
        LIMIT 50
    """).fetchall()

    vkupno = cursor.execute("""
        SELECT p.part_number, p.ime,
               SUM(CASE WHEN d.plateno=1 THEN d.kolicina ELSE 0 END) AS platena,
               SUM(CASE WHEN d.plateno=0 THEN d.kolicina ELSE 0 END) AS neplatena,
               COUNT(*) AS redovi
        FROM zaliha_dodadi d
        LEFT JOIN parts p ON d.artikl_id = p.id
        GROUP BY d.artikl_id
    """).fetchall()

    conn.close()

    html = "<h2>zaliha_dodadi (последни 50)</h2><table border=1 cellpadding=5>"
    html += "<tr><th>ID</th><th>artikl_id</th><th>PN</th><th>Ime</th><th>Kolicina</th><th>Plateno</th><th>Datum</th><th>Username</th></tr>"
    for r in site:
        r = dict(r)
        html += f"<tr><td>{r['id']}</td><td>{r['artikl_id']}</td><td>{r['part_number']}</td><td>{r['ime']}</td><td>{r['kolicina']}</td><td>{r['plateno']}</td><td>{r['datum']}</td><td>{r['username']}</td></tr>"
    html += "</table>"

    html += "<br><h2>Вкупно по артикл</h2><table border=1 cellpadding=5>"
    html += "<tr><th>PN</th><th>Ime</th><th>Platena</th><th>Neplatena</th><th>Redovi</th></tr>"
    for r in vkupno:
        r = dict(r)
        html += f"<tr><td>{r['part_number']}</td><td>{r['ime']}</td><td>{r['platena']}</td><td>{r['neplatena']}</td><td>{r['redovi']}</td></tr>"
    html += "</table>"

    return html


# ─────────────────────────────────────────────────────────────
# PENDING ИЗВОЗИ
# ─────────────────────────────────────────────────────────────
@zalihi_bp.route("/pending")
@login_required
def pending():
    try:
        pending_items = get_pending_exports()
        return render_template("zalihi_pending.html", pending_items=pending_items)
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Грешка: {str(e)}", "danger")
        return render_template("zalihi_pending.html", pending_items=[])

@zalihi_bp.route("/pending/odobri", methods=["POST"])
@login_required
def pending_odobri():
    data = request.get_json() or {}
    try:
        pending_id = int(data.get("id", 0))
    except (TypeError, ValueError):
        pending_id = 0
    if pending_id <= 0:
        return jsonify({"success": False, "message": "Невалиден ID"}), 400
    response, status_code = approve_pending_export(pending_id, session["user"])
    return jsonify(response), status_code

@zalihi_bp.route("/pending/odbij", methods=["POST"])
@login_required
def pending_odbij():
    data = request.get_json() or {}
    try:
        pending_id = int(data.get("id", 0))
    except (TypeError, ValueError):
        pending_id = 0
    if pending_id <= 0:
        return jsonify({"success": False, "message": "Невалиден ID"}), 400
    response, status_code = reject_pending_export(pending_id, session["user"])
    return jsonify(response), status_code

@zalihi_bp.route("/pending/odobri_site", methods=["POST"])
@login_required
def pending_odobri_site():
    response, status_code = approve_all_pending_exports(session["user"])
    return jsonify(response), status_code

@zalihi_bp.route("/dodadeni_po_nedeli")
@login_required
def dodadeni_po_nedeli():
    conn = get_db()
    cursor = conn.cursor()
    raw = cursor.execute("""
        SELECT
            datum,
            pn,
            ime,
            SUM(kolicina) AS vkupna_kolicina,
            GROUP_CONCAT(username, ', ') AS korisnici_raw,
            zabeleska
        FROM zaliha_dodadi_log
        GROUP BY pn, ime, strftime('%Y-%W', datum)
        ORDER BY datum DESC
    """).fetchall()
    conn.close()
    weeks = defaultdict(list)
    week_totals = {}
    for row in raw:
        r = dict(row)
        try:
            dt = datetime.strptime(r['datum'][:10], '%Y-%m-%d')
            week_num = dt.isocalendar()[1]
            nedela_key = f"КН{week_num:02d}"
        except:
            nedela_key = "КН--"
        r['nedela'] = nedela_key
        if r.get('korisnici_raw'):
            users = sorted(set(u.strip() for u in r['korisnici_raw'].split(',')))
            r['korisnici'] = ', '.join(users)
        else:
            r['korisnici'] = '—'
        weeks[nedela_key].append(r)
        if nedela_key not in week_totals:
            week_totals[nedela_key] = 0
        week_totals[nedela_key] += r['vkupna_kolicina']
    sorted_weeks = sorted(weeks.items(), key=lambda x: x[0], reverse=True)
    return render_template('zalihi_dodadeni_po_nedeli.html',
                           sorted_weeks=sorted_weeks,
                           week_totals=week_totals)


@zalihi_bp.route("/dodadeni_po_nedeli/export")
@login_required
def dodadeni_po_nedeli_export():
    conn = get_db()
    cursor = conn.cursor()
    data = cursor.execute("""
        SELECT
            datum,
            pn,
            ime,
            SUM(kolicina) AS vkupna_kolicina,
            GROUP_CONCAT(username, ', ') AS korisnici_raw,
            zabeleska
        FROM zaliha_dodadi_log
        GROUP BY pn, ime
        ORDER BY datum DESC
    """).fetchall()
    conn.close()
    cleaned_data = prepare_dodadeni_po_nedeli_export_rows(data)
    output = build_dodadeni_po_nedeli_workbook(cleaned_data)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"zaliha_po_nedeli_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )

@zalihi_bp.route("/istorija")
@login_required
def istorija():
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT datum, username, pn, ime, kolicina, tip
        FROM zaliha_izvoz_log
        ORDER BY datum DESC
        LIMIT 200
    """).fetchall()
    conn.close()
    logovi = prepare_istorija_rows(rows)
    return render_template('zalihi_istorija.html', logovi=logovi)

@zalihi_bp.route("/izvozi_po_nedeli")
@login_required
def izvozi_po_nedeli():
    conn = get_db()
    cursor = conn.cursor()

    rows = cursor.execute("""
        SELECT datum, username, pn, ime, kolicina, tip
        FROM zaliha_izvoz_log
        ORDER BY datum ASC
    """).fetchall()
    conn.close()

    week_data = group_izvozi_po_nedeli(rows)
    sorted_weeks, week_totals, week_platena, week_neplatena = build_izvozi_po_nedeli_view_model(week_data)

    return render_template(
        'zalihi_izvozi_po_nedeli.html',
        sorted_weeks=sorted_weeks,
        week_totals=week_totals,
        week_platena=week_platena,
        week_neplatena=week_neplatena
    )

@zalihi_bp.route("/izvozi_po_nedeli/export")
@login_required
def izvozi_po_nedeli_export():
    conn = get_db()
    cursor = conn.cursor()

    rows = cursor.execute("""
        SELECT datum, username, pn, ime, kolicina, tip
        FROM zaliha_izvoz_log
        ORDER BY datum ASC
    """).fetchall()
    conn.close()

    week_data = group_izvozi_po_nedeli(rows)
    output = build_izvozi_po_nedeli_workbook(week_data)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"izvozi_po_nedeli_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )

@zalihi_bp.route("/istorija/export")
@login_required
def istorija_export():
    conn = get_db()
    cursor = conn.cursor()
    rows = cursor.execute("""
        SELECT datum, username, pn, ime, kolicina, tip
        FROM zaliha_izvoz_log
        ORDER BY datum DESC
    """).fetchall()
    conn.close()

    output = build_istorija_workbook(prepare_istorija_rows(rows))
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"istorija_izvozi_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )

@zalihi_bp.route("/profakturi/nova", methods=["GET", "POST"])
@login_required
def profakturi_nova():
    if request.method == "POST":
        response, status_code = create_profaktura(request.get_json() or {}, session["user"])
        return jsonify(response), status_code

    artikli = get_profakturi_nova_artikli()
    return render_template("profakturi_nova.html", artikli=artikli, today=date.today().isoformat())

@zalihi_bp.route("/profakturi/pending")
@login_required
def profakturi_pending():
    profakturi = get_pending_profakturi()
    return render_template("profakturi_pending.html", profakturi=profakturi)

@zalihi_bp.route("/profakturi/odobri", methods=["POST"])
@login_required
def profakturi_odobri():
    data = request.get_json() or {}
    profaktura_id = int(data.get("id", 0))
    if profaktura_id <= 0:
        return jsonify({"success": False, "message": "Невалиден ID"}), 400

    response, status_code = approve_profaktura(profaktura_id, session["user"])
    return jsonify(response), status_code

@zalihi_bp.route("/profakturi/odbij", methods=["POST"])
@login_required
def profakturi_odbij():
    data = request.get_json() or {}
    profaktura_id = int(data.get("id", 0))
    if profaktura_id <= 0:
        return jsonify({"success": False, "message": "Невалиден ID"}), 400

    response, status_code = reject_profaktura(profaktura_id, session["user"])
    return jsonify(response), status_code

@zalihi_bp.route("/storno/zaliha/<int:artikl_id>")
@login_required
def storno_zaliha(artikl_id):
    """Враќа JSON со тековни залихи (платена / неплатена) за дадениот артикл."""
    try:
        conn = get_db()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        row = cursor.execute("""
            SELECT
                COALESCE(SUM(CASE WHEN plateno=1 THEN kolicina ELSE 0 END), 0) AS platena,
                COALESCE(SUM(CASE WHEN plateno=0 THEN kolicina ELSE 0 END), 0) AS neplatena
            FROM zaliha_dodadi
            WHERE artikl_id = ?
        """, (artikl_id,)).fetchone()

        conn.close()
        return jsonify({
            'platena':   int(row['platena'])   if row else 0,
            'neplatena': int(row['neplatena']) if row else 0,
        })
    except Exception as e:
        return jsonify({'platena': 0, 'neplatena': 0, 'error': str(e)}), 500


@zalihi_bp.route("/storno", methods=["GET", "POST"])
@login_required
def storno():
    if request.method == "POST":
        response, status_code = execute_storno(request.get_json() or {}, session["user"])
        return jsonify(response), status_code

    try:
        storna_list = [dict(s) for s in get_storno_history()]
    except Exception as e:
        print(f"Грешка при читање сторно историја: {e}")
        storna_list = []

    return render_template("zalihi_storno.html", storna=storna_list)
