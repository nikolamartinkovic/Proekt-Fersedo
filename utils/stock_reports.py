import os
import smtplib
import traceback
from datetime import datetime
from email.utils import formataddr
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO

from flask import current_app
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from utils.db import get_db


def _fetch_stock_rows():
    conn = get_db()
    cursor = conn.cursor()
    rows = [
        dict(r)
        for r in cursor.execute(
            """
            SELECT p.part_number, p.ime, COALESCE(SUM(d.kolicina), 0) AS total
            FROM parts p LEFT JOIN zaliha_dodadi d ON p.id = d.artikl_id
            GROUP BY p.id HAVING total > 0 ORDER BY p.part_number
            """
        ).fetchall()
    ]
    conn.close()
    return rows


def generiraj_zaliha_excel():
    try:
        rows = _fetch_stock_rows()
        static_folder = current_app.config["STATIC_FOLDER"]

        wb = Workbook()
        ws = wb.active
        ws.title = "Current Stock"

        logo_path = os.path.join(static_folder, "logo2.png")
        ws.merge_cells("A1:C7")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        if os.path.exists(logo_path):
            try:
                img = OpenpyxlImage(logo_path)
                img.width = 280
                img.height = 110
                ws.add_image(img, "B1")
            except Exception as e:
                print(f"[LOGO ERROR] {e}")
        else:
            ws["A1"].value = "Fersedo"
            ws["A1"].font = Font(name="Calibri", bold=True, size=32, color="1E3A8A")

        ws.row_dimensions[8].height = 25
        title_row = 9
        ws.merge_cells(f"A{title_row}:C{title_row}")
        tc = ws[f"A{title_row}"]
        tc.value = f"CURRENT STOCK - {datetime.now().strftime('%d-%m-%Y')}"
        tc.font = Font(name="Calibri", bold=True, size=18, color="1E40AF")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[title_row].height = 40

        header_row = title_row + 2
        hfont = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
        hfill = PatternFill("solid", fgColor="2563EB")
        border = Border(**{s: Side(style="thin", color="D1D5DB") for s in ("left", "right", "top", "bottom")})
        for col_idx, header in enumerate(["Part Number", "Description", "Total Quantity"], 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        ws.row_dimensions[header_row].height = 36
        data_start = header_row + 1
        thin_border = Border(**{s: Side(style="thin", color="E5E7EB") for s in ("left", "right", "top", "bottom")})
        for i, row in enumerate(rows):
            r = data_start + i
            fill = PatternFill("solid", fgColor="F9FAFB" if i % 2 == 0 else "FFFFFF")
            ws.cell(row=r, column=1).value = row["part_number"]
            ws.cell(row=r, column=2).value = row["ime"] or "Unknown item"
            ws.cell(row=r, column=3).value = row["total"]
            for col in range(1, 4):
                cell = ws.cell(row=r, column=col)
                cell.fill = fill
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=11)
                cell.alignment = Alignment(horizontal="center" if col == 3 else "left", vertical="center")

        total_row = data_start + len(rows) + 2
        ws.cell(row=total_row, column=1).value = "TOTAL"
        ws.cell(row=total_row, column=3).value = sum(r["total"] for r in rows)
        tfont = Font(name="Calibri", bold=True, size=14, color="065F46")
        tfill = PatternFill("solid", fgColor="D1FAE5")
        tborder = Border(**{s: Side(style="medium", color="10B981") for s in ("left", "right", "top", "bottom")})
        for col in [1, 3]:
            cell = ws.cell(row=total_row, column=col)
            cell.font = tfont
            cell.fill = tfill
            cell.border = tborder
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[total_row].height = 42
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 20
        ws.freeze_panes = f"A{header_row + 1}"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        print(f"[EMAIL ZALIHA] Excel generated - {len(rows)} items")
        return output
    except Exception as e:
        print(f"[EMAIL ZALIHA] Error: {e}")
        traceback.print_exc()
        return None


def generiraj_zaliha_pdf():
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=20 * mm,
        leftMargin=20 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=18, textColor=colors.darkblue, spaceAfter=12)
    logo_path = os.path.join(current_app.config["STATIC_FOLDER"], "logo2.png")
    if os.path.exists(logo_path):
        try:
            logo_img = Image(logo_path, width=120 * mm, height=40 * mm)
            logo_img.hAlign = "CENTER"
            elements.append(logo_img)
            elements.append(Spacer(1, 8 * mm))
        except Exception as e:
            print(f"[PDF LOGO ERROR] {e}")

        elements.append(Paragraph(f"Current Stock Report - {datetime.now().strftime('%d-%m-%Y')}", title_style))
    elements.append(Spacer(1, 12 * mm))

    table_data = [["Part Number", "Description", "Total Quantity"]]
    total_qty = 0
    for row in _fetch_stock_rows():
        table_data.append([row["part_number"], row["ime"] or "Unknown item", row["total"]])
        total_qty += row["total"]
    table_data.append(["TOTAL", "", total_qty])

    table = Table(table_data, colWidths=[50 * mm, 90 * mm, 40 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (-1, -1), (-1, -1), colors.HexColor("#D1FAE5")),
                ("FONTNAME", (-1, -1), (-1, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F9FAFB")]),
            ]
        )
    )
    elements.append(table)
    elements.append(Spacer(1, 20 * mm))
    elements.append(
        Paragraph(
            f"Generated by Fersedo Production System | {datetime.now().strftime('%d-%m-%Y %H:%M')}",
            styles["Normal"],
        )
    )

    doc.build(elements)
    buffer.seek(0)
    return buffer


def isprati_zaliha_email():
    print(f"[EMAIL ZALIHA] Почнува - {datetime.now().strftime('%d-%m-%Y %H:%M')}")
    try:
        conn = get_db()
        rows = conn.execute("SELECT email, tip FROM email_recipients WHERE aktiven = 1").fetchall()
        conn.close()

        to_list = [r["email"] for r in rows if r["tip"] == "to"]
        cc_list = [r["email"] for r in rows if r["tip"] == "cc"]
        if not to_list:
            print("[EMAIL ZALIHA] Нема To примачи. Прескокнување.")
            return

        excel_buffer = generiraj_zaliha_excel()
        pdf_buffer = generiraj_zaliha_pdf()
        if not excel_buffer or not pdf_buffer:
            print("[EMAIL ZALIHA] Неуспешно генерирање на прилог(и)")
            return

        datum_str = datetime.now().strftime("%d-%m-%Y")
        excel_filename = f"stock_report_{datum_str.replace('.', '')}.xlsx"
        pdf_filename = f"stock_report_{datum_str.replace('.', '')}.pdf"

        msg = MIMEMultipart()
        msg["From"] = formataddr(
            (
                current_app.config.get("EMAIL_FROM_NAME", "Info Fersedo"),
                current_app.config["EMAIL_HOST_USER"],
            )
        )
        msg["To"] = ", ".join(to_list)
        if cc_list:
            msg["Cc"] = ", ".join(cc_list)
        msg["Subject"] = f"Current Stock Report - {datum_str}"
        msg.attach(
            MIMEText(
                (
                    f"Dear team,\nAttached are the current Fersedo stock reports as of {datum_str}.\n"
                    "This is an automatically generated report.\nBest regards,\nFersedo System"
                ),
                "plain",
                "utf-8",
            )
        )

        for buf, fname in [(excel_buffer, excel_filename), (pdf_buffer, pdf_filename)]:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(buf.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{fname}"')
            msg.attach(part)

        with smtplib.SMTP(current_app.config["EMAIL_HOST"], current_app.config["EMAIL_PORT"]) as server:
            server.starttls()
            server.login(current_app.config["EMAIL_HOST_USER"], current_app.config["EMAIL_HOST_PASSWORD"])
            server.sendmail(current_app.config["EMAIL_HOST_USER"], to_list + cc_list, msg.as_string())
        print(f"[EMAIL ZALIHA] Успешно испратено - To: {to_list}, Cc: {cc_list}")
    except Exception as e:
        print(f"[EMAIL ZALIHA] Грешка: {e}")
        traceback.print_exc()
