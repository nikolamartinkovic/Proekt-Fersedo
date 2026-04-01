from collections import defaultdict
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def get_nedela(datum_str):
    try:
        dt = datetime.strptime(str(datum_str)[:10], "%Y-%m-%d")
        week = dt.isocalendar()[1]
        return f"КН{week:02d}"
    except Exception:
        return "КН--"

def _row_to_dict(row, keys):
    if hasattr(row, "keys"):
        return dict(row)
    return {key: row[idx] for idx, key in enumerate(keys)}


def _autosize_columns(ws, extra=4):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + extra


def prepare_dodadeni_po_nedeli_export_rows(rows):
    cleaned_data = []
    for row in rows:
        data = dict(row)
        data["nedela"] = get_nedela(data["datum"])
        if data.get("korisnici_raw"):
            users = sorted(set(user.strip() for user in data["korisnici_raw"].split(",")))
            data["korisnici"] = ", ".join(users)
        else:
            data["korisnici"] = "—"
        cleaned_data.append(data)
    return cleaned_data


def build_dodadeni_po_nedeli_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Zaliha Po Nedeli"
    ws.append(["Датум", "Недела", "PN", "Име на артикл", "Вкупна количина", "Корисници", "Забелешка"])
    for row in rows:
        ws.append(
            [
                row["datum"][:16],
                row["nedela"],
                row["pn"],
                row["ime"],
                row["vkupna_kolicina"],
                row["korisnici"],
                row["zabeleska"] or "—",
            ]
        )
    _autosize_columns(ws, extra=5)
    return workbook_to_bytes(wb)


def prepare_istorija_rows(rows):
    prepared = []
    keys = ["datum", "username", "pn", "ime", "kolicina", "tip"]
    for row in rows:
        data = _row_to_dict(row, keys)
        data["nedela"] = get_nedela(data["datum"])
        prepared.append(data)
    return prepared


def group_izvozi_po_nedeli(rows):
    week_data = defaultdict(
        lambda: defaultdict(
            lambda: {"ime": "—", "kolicina_platena": 0, "kolicina_neplatena": 0, "korisnici": set()}
        )
    )
    keys = ["datum", "username", "pn", "ime", "kolicina", "tip"]

    for row in rows:
        data = _row_to_dict(row, keys)
        try:
            dt = datetime.strptime(str(data["datum"])[:10], "%Y-%m-%d")
            week_num = dt.isocalendar()[1]
            year = dt.isocalendar()[0]
            week_key = f"{year}-КН{week_num:02d}"
        except Exception:
            week_key = "----КН--"

        pn = data["pn"]
        entry = week_data[week_key][pn]
        entry["ime"] = data["ime"] or "—"
        entry["korisnici"].add(data["username"])

        if "Платена" in (data["tip"] or ""):
            entry["kolicina_platena"] += int(data["kolicina"])
        else:
            entry["kolicina_neplatena"] += int(data["kolicina"])

    return week_data


def build_izvozi_po_nedeli_view_model(week_data):
    week_totals = {}
    week_platena = {}
    week_neplatena = {}
    sorted_weeks = []

    for week_key in sorted(week_data.keys(), reverse=True):
        pn_dict = week_data[week_key]
        artikli = []
        total = 0
        total_platena = 0
        total_neplatena = 0

        for pn, entry in sorted(pn_dict.items()):
            vkupno = entry["kolicina_platena"] + entry["kolicina_neplatena"]
            artikli.append(
                {
                    "pn": pn,
                    "ime": entry["ime"],
                    "kolicina_platena": entry["kolicina_platena"],
                    "kolicina_neplatena": entry["kolicina_neplatena"],
                    "vkupno": vkupno,
                    "korisnici": ", ".join(sorted(entry["korisnici"])),
                }
            )
            total += vkupno
            total_platena += entry["kolicina_platena"]
            total_neplatena += entry["kolicina_neplatena"]

        display_key = week_key.split("-", 1)[1] if "-" in week_key else week_key
        sorted_weeks.append((display_key, artikli))
        week_totals[display_key] = total
        week_platena[display_key] = total_platena
        week_neplatena[display_key] = total_neplatena

    return sorted_weeks, week_totals, week_platena, week_neplatena


def build_izvozi_po_nedeli_workbook(week_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Izvozi Po Nedeli"

    header_fill = PatternFill("solid", fgColor="1e40af")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    total_fill = PatternFill("solid", fgColor="bbf7d0")
    total_font = Font(bold=True)

    headers = ["Недела", "PN", "Име на артикл", "Платена", "Неплатена", "Вкупно", "Корисници"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_num = 2
    for week_key in sorted(week_data.keys(), reverse=True):
        display_key = week_key.split("-", 1)[1] if "-" in week_key else week_key
        pn_dict = week_data[week_key]
        total = 0
        total_platena = 0
        total_neplatena = 0

        for pn, entry in sorted(pn_dict.items()):
            vkupno = entry["kolicina_platena"] + entry["kolicina_neplatena"]
            ws.append(
                [
                    display_key,
                    pn,
                    entry["ime"],
                    entry["kolicina_platena"],
                    entry["kolicina_neplatena"],
                    vkupno,
                    ", ".join(sorted(entry["korisnici"])),
                ]
            )
            total += vkupno
            total_platena += entry["kolicina_platena"]
            total_neplatena += entry["kolicina_neplatena"]
            row_num += 1

        ws.append([f"ВКУПНО {display_key}", "", "", total_platena, total_neplatena, total, ""])
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = total_fill
            cell.font = total_font
        row_num += 2
        ws.append([])

    _autosize_columns(ws)
    return workbook_to_bytes(wb)


def build_istorija_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Istorija Izvozi"
    ws.append(["Датум", "Корисник", "PN", "Име на артикл", "Количина", "Тип"])

    for row in rows:
        ws.append(
            [
                row["datum"][:16],
                row["username"],
                row["pn"],
                row["ime"],
                row["kolicina"],
                row["tip"],
            ]
        )

    _autosize_columns(ws, extra=5)
    return workbook_to_bytes(wb)


def workbook_to_bytes(workbook):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
